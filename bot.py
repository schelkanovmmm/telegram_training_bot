import os
import sqlite3
import logging
import re
import httpx
import csv
import io
import secrets
import asyncio
import xml.etree.ElementTree as ET
from aiohttp import web
from datetime import datetime, timedelta
from typing import Optional, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference

from telegram import Update, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

DB_PATH = os.getenv("DB_PATH", "training_bot.db")
BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN","8653934626:AAFdgvBp4R-ads4Q4uSzgVKeOf8bljdtCtc")

SESSION_DATE, CUSTOM_EXERCISE_NAME, SESSION_QUICK_INPUT, SESSION_MANUAL_INPUT, SESSION_RPE, SESSION_NOTES, SESSION_EDIT, AI_CHAT, MEASURE_INPUT, NUTRITION_INPUT, IMPORT_SCALE, IMPORT_HEALTH = range(12)

PROGRAMS = {
    # ─── FULL BODY A — СИЛА + V-SHAPE ──────────────────────────────────────────
    # Верх получает силовой стимул, спина расширяется, ноги просто поддерживаются
    "A": {
        "title": "FULL BODY A — СИЛА + V-SHAPE 💪",
        "notes": [
            "Акцент: сила верха + V-shape",
            "Отдых: 2–3 мин на базовых, 60–90 сек на изоляции",
        ],
        "review_weeks": 10,
        "exercises": [
            {
                # 1. Горизонтальный жим — главная силовая база верха
                "group": "Грудь", "exercise": "Жим штанги лёжа",
                "sets": 3, "reps": "5–6", "rpe": "RIR 1–2", "step": "+2.5 кг",
                "rest": "2–3 мин",
                "desc": (
                    "💪 ЖИМ ШТАНГИ ЛЁЖА\n"
                    "Категория: горизонтальный жим\n"
                    "Цель: главная силовая база верха\n\n"
                    "Техника:\n"
                    "• Лопатки сведены и прижаты к скамье\n"
                    "• Хват чуть шире плеч, запястья прямые\n"
                    "• Опускай до касания груди — без отскока\n"
                    "• Жим мощно вверх, локти ~75° к корпусу\n"
                    "• Ноги на полу, поясница нейтральная\n\n"
                    "RIR 1–2: останавливайся за 1–2 повтора до отказа\n"
                    "Прогрессия: +2.5 кг когда все 3×6 с RIR 1–2"
                )
            },
            {
                # 2. Вертикальная тяга — приоритет на ширину спины
                "group": "Спина", "exercise": "Подтягивания с весом",
                "sets": 3, "reps": "6–8", "rpe": "RIR 1–2", "step": "+1.25–2.5 кг",
                "rest": "2 мин",
                "desc": (
                    "💪 ПОДТЯГИВАНИЯ С ВЕСОМ\n"
                    "Категория: вертикальная тяга\n"
                    "Цель: ширина спины, V-shape\n\n"
                    "Нет возможности? → тяга верхнего блока\n\n"
                    "Техника:\n"
                    "• Широкий хват, ладони от себя\n"
                    "• Тянись грудью к перекладине, локти вниз-назад\n"
                    "• Полная амплитуда: вниз до выпрямления рук\n"
                    "• Без раскачки, движение контролируемое\n\n"
                    "RIR 1–2: без отказа\n"
                    "Прогрессия: +1.25–2.5 кг к поясу"
                )
            },
            {
                # 3. Движение ног — поддержка, не приоритет
                "group": "Ноги", "exercise": "Жим ногами",
                "sets": 2, "reps": "6–8", "rpe": "RIR 2", "step": "+5–10 кг",
                "rest": "2 мин",
                "desc": (
                    "💪 ЖИМ НОГАМИ\n"
                    "Категория: движение ног\n"
                    "Цель: поддержка нижней части — НЕ приоритет\n\n"
                    "Альтернатива: присед со штангой (если удобно)\n\n"
                    "Техника:\n"
                    "• Стопы на ширине плеч, середина-верх платформы\n"
                    "• Угол 90° в коленях в нижней точке\n"
                    "• Не отрывай поясницу от спинки\n"
                    "• Держи напряжение — не до замка вверху\n\n"
                    "RIR 2: 2 повтора в запасе — не убиваться!\n"
                    "Прогрессия: +5–10 кг"
                )
            },
            {
                # 4. Тазовое движение — задняя цепь без перегруза поясницы
                "group": "Задняя цепь", "exercise": "Сгибание ног лёжа",
                "sets": 2, "reps": "10–12", "rpe": "RIR 2–3", "step": "след. стек",
                "rest": "90 сек",
                "desc": (
                    "💪 СГИБАНИЕ НОГ ЛЁЖА\n"
                    "Категория: тазовое движение\n"
                    "Цель: задняя цепь без нагрузки на поясницу\n\n"
                    "Альтернатива: лёгкая румынская тяга (RIR 3)\n\n"
                    "Техника:\n"
                    "• Лёжа на животе, валик над пятками\n"
                    "• Сгибай до 90° или дальше\n"
                    "• Пик сокращения — задержка 1 сек\n"
                    "• Медленный эксцентрик 2–3 сек\n"
                    "• Бёдра прижаты к скамье\n\n"
                    "RIR 2–3: легко, техника важнее веса\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 5. Горизонтальная тяга — толщина спины без нагрузки на поясницу
                "group": "Спина", "exercise": "Тяга с упором грудью (Chest-supported row)",
                "sets": 3, "reps": "8–10", "rpe": "RIR 1–2", "step": "+1–2 кг",
                "rest": "90 сек",
                "desc": (
                    "💪 ТЯГА С УПОРОМ ГРУДЬЮ\n"
                    "Категория: горизонтальная тяга\n"
                    "Цель: толщина спины без нагрузки на поясницу\n\n"
                    "Варианты:\n"
                    "• Тяга гантелей с упором грудью на наклонной скамье\n"
                    "• Chest-supported row машина (если есть в зале)\n"
                    "• Тяга горизонтального блока (если нет выбора)\n\n"
                    "Техника:\n"
                    "• Грудь на скамье — поясница полностью выключена\n"
                    "• Тяни к нижней части груди, локти назад\n"
                    "• Сводить лопатки в конечной точке — задержка 1 сек\n"
                    "• Полная амплитуда вниз — растяжение лопаток\n\n"
                    "RIR 1–2\n"
                    "Прогрессия: +1–2 кг на гантель"
                )
            },
            {
                # 6. Плечи — главный визуальный блок
                "group": "Плечи", "exercise": "Разведения гантелей в стороны",
                "sets": 4, "reps": "12–20", "rpe": "RIR 1", "step": "+0.5–1 кг",
                "rest": "60–75 сек",
                "desc": (
                    "💪 РАЗВЕДЕНИЯ ГАНТЕЛЕЙ В СТОРОНЫ\n"
                    "Категория: плечи\n"
                    "Цель: средняя дельта — главный визуальный блок\n\n"
                    "Техника:\n"
                    "• Стоя, лёгкий наклон вперёд (~10–15°)\n"
                    "• Веди движение локтём, не кистью\n"
                    "• Поднимай до уровня плеч — не выше\n"
                    "• Задержка 1 сек наверху\n"
                    "• Медленный эксцентрик 2–3 сек вниз\n"
                    "• Без читинга корпусом!\n\n"
                    "Прогрессия: сначала добирай повторы (12→15→18→20),\n"
                    "потом повышай вес (+0.5–1 кг)\n"
                    "RIR 1: жжение — норма"
                )
            },
            {
                # 7. Задняя дельта — здоровье плеч
                "group": "Задняя дельта", "exercise": "Face Pull",
                "sets": 3, "reps": "15–20", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60 сек",
                "desc": (
                    "💪 FACE PULL\n"
                    "Категория: задняя дельта\n"
                    "Цель: задняя дельта, здоровье плеч, осанка\n\n"
                    "Альтернатива: Reverse Pec Deck\n\n"
                    "Техника:\n"
                    "• Блок на уровне лица или выше, канат с двумя ручками\n"
                    "• Тяни к лицу, локти выше плеч — в стороны\n"
                    "• В конечной точке — внешняя ротация плеча\n"
                    "• Медленно возвращай, не теряй напряжение\n\n"
                    "RIR 1: работай на жжение\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 8. Кор — контроль, без раскачки
                "group": "Пресс", "exercise": "Подъём ног в висе",
                "sets": 3, "reps": "10–15", "rpe": "RIR 1–2", "step": "+1–2 повт",
                "rest": "60–90 сек",
                "desc": (
                    "💪 ПОДЪЁМ НОГ В ВИСЕ\n"
                    "Категория: кор\n"
                    "Цель: нижний пресс, контроль без раскачки\n\n"
                    "Альтернатива: Cable Crunch\n\n"
                    "Техника:\n"
                    "• Вис на перекладине, тело стабильно\n"
                    "• Поднимай прямые ноги — без раскачки!\n"
                    "• В верхней точке — задержка 1 сек\n"
                    "• Медленно опускай 2–3 сек\n\n"
                    "RIR 1–2: контроль важнее количества\n"
                    "Прогрессия: +1–2 повтора"
                )
            },
        ],
    },

    # ─── FULL BODY B — ВЕРХ ГРУДИ + ДЕЛЬТЫ + ШИРОЧАЙШИЕ ───────────────────────
    # Главный визуальный день: верх груди, дельты, широчайшие
    # Ноги есть, но не забирают восстановление
    "B": {
        "title": "FULL BODY B — ВЕРХ ГРУДИ + ДЕЛЬТЫ 🔥",
        "notes": [
            "Акцент: верх груди + дельты + широчайшие",
            "Отдых: 90 сек на базовых, 60 сек на изоляции",
        ],
        "review_weeks": 10,
        "exercises": [
            {
                # 1. Горизонтальный жим — приоритет верх груди
                "group": "Грудь", "exercise": "Жим гантелей на наклонной скамье",
                "sets": 4, "reps": "8–10", "rpe": "RIR 1–2", "step": "+1–2 кг",
                "rest": "90 сек",
                "desc": (
                    "🔥 ЖИМ ГАНТЕЛЕЙ НА НАКЛОННОЙ\n"
                    "Категория: горизонтальный жим\n"
                    "Цель: приоритет верх груди\n\n"
                    "Техника:\n"
                    "• Угол скамьи 30–45°\n"
                    "• Гантели на уровне верхней части груди\n"
                    "• Локти под углом 45–75° к корпусу\n"
                    "• Полная амплитуда — растяжение груди внизу\n"
                    "• Жим вверх и чуть внутрь — пик сокращения\n\n"
                    "RIR 1–2: 4 тяжёлых рабочих подхода\n"
                    "Прогрессия: +1–2 кг"
                )
            },
            {
                # 2. Вертикальная тяга — широчайшие, V-shape
                "group": "Спина", "exercise": "Тяга верхнего блока нейтральным хватом",
                "sets": 4, "reps": "8–12", "rpe": "RIR 1–2", "step": "след. стек",
                "rest": "90 сек",
                "desc": (
                    "🔥 ТЯГА ВЕРХНЕГО БЛОКА НЕЙТРАЛЬНЫМ ХВАТОМ\n"
                    "Категория: вертикальная тяга\n"
                    "Цель: широчайшие, V-shape\n\n"
                    "Нейтральный или узкий хват (параллельная рукоять)\n\n"
                    "Техника:\n"
                    "• Бёдра зафиксированы валиками\n"
                    "• Небольшой наклон назад — норма\n"
                    "• Тяни к верхней части груди, локти вниз-назад\n"
                    "• Сводить лопатки в нижней точке — задержка 1 сек\n"
                    "• Медленный возврат — растяжение наверху\n\n"
                    "RIR 1–2: 3–4 рабочих подхода\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 3. Движение ног — поддержка, без отказа
                "group": "Ноги", "exercise": "Болгарские выпады",
                "sets": 2, "reps": "10 на ногу", "rpe": "RIR 2", "step": "+1–2 кг",
                "rest": "90 сек",
                "desc": (
                    "🔥 БОЛГАРСКИЕ ВЫПАДЫ\n"
                    "Категория: движение ног\n"
                    "Цель: ноги поддержка — без отказа!\n\n"
                    "Техника:\n"
                    "• Задняя нога на скамье, передняя шаг вперёд\n"
                    "• Опускайся вертикально вниз\n"
                    "• Переднее колено не выходит за носок\n"
                    "• Корпус прямой, не наклоняй вперёд\n"
                    "• Сначала все повторы на одну ногу, потом другую\n\n"
                    "RIR 2: 2 повтора в запасе\n"
                    "Прогрессия: +1–2 кг"
                )
            },
            {
                # 4. Тазовое движение — мягче чем тяжёлая тяга
                "group": "Задняя цепь", "exercise": "Pull-through в кроссовере",
                "sets": 3, "reps": "10–15", "rpe": "RIR 1–2", "step": "след. стек",
                "rest": "75–90 сек",
                "desc": (
                    "🔥 PULL-THROUGH В КРОССОВЕРЕ\n"
                    "Категория: тазовое движение\n"
                    "Цель: задняя цепь — мягче чем тяжёлая тяга\n\n"
                    "Альтернатива: сгибание ног лёжа\n\n"
                    "Техника Pull-through:\n"
                    "• Блок внизу за спиной, канат между ног\n"
                    "• Наклон вперёд с прямой спиной, таз назад\n"
                    "• Разгибание через бедро — тяни тазом вперёд\n"
                    "• Акцент на ягодицы и хамстринги, не на руки\n\n"
                    "RIR 1–2: мягче чем тяжёлая тяга\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 5. Горизонтальная тяга — амплитуда, выравнивание сторон
                "group": "Спина", "exercise": "Одноручная тяга в кроссовере",
                "sets": 3, "reps": "10–12", "rpe": "RIR 1–2", "step": "след. стек",
                "rest": "90 сек",
                "desc": (
                    "🔥 ОДНОРУЧНАЯ ТЯГА В КРОССОВЕРЕ\n"
                    "Категория: горизонтальная тяга\n"
                    "Цель: широчайшие, амплитуда, выравнивание сторон\n\n"
                    "Альтернатива: тяга гантели одной рукой с упором\n\n"
                    "Техника:\n"
                    "• Блок на уровне пояса или выше\n"
                    "• Тяни к бедру/поясу, локоть назад\n"
                    "• Полная амплитуда — растяжение в начале\n"
                    "• Сначала слабая сторона, потом сильная\n"
                    "• Одинаковое количество повторов на обе стороны\n\n"
                    "RIR 1–2: исправляет асимметрию спины\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 6. Плечи — главный эстетический акцент
                "group": "Плечи", "exercise": "Cable Lateral Raise",
                "sets": 4, "reps": "12–20", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60 сек",
                "desc": (
                    "🔥 CABLE LATERAL RAISE\n"
                    "Категория: плечи\n"
                    "Цель: средняя дельта — главный эстетический акцент\n\n"
                    "Преимущество перед гантелями: постоянное натяжение\n\n"
                    "Техника:\n"
                    "• Блок внизу, стоя сбоку\n"
                    "• Рука слегка согнута в локте\n"
                    "• Поднимай до уровня плеча дугообразно\n"
                    "• Медленно опускай 2–3 сек\n\n"
                    "Прогрессия: сначала добирай повторы (12→20),\n"
                    "потом следующий стек\n"
                    "RIR 1: жжение — цель"
                )
            },
            {
                # 7. Грудь добивка — верх груди
                "group": "Грудь", "exercise": "Cable Fly (low-to-high)",
                "sets": 2, "reps": "12–15", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60 сек",
                "desc": (
                    "🔥 CABLE FLY LOW-TO-HIGH\n"
                    "Категория: грудь добивка\n"
                    "Цель: верх груди, внутренняя грудь\n\n"
                    "Техника:\n"
                    "• Блоки внизу (у пола), стоя в центре\n"
                    "• Движение снизу-вверх и к центру — дугой\n"
                    "• В верхней точке скрести руки — пик сокращения\n"
                    "• Руки слегка согнуты в локтях — угол не меняется\n\n"
                    "RIR 1: памп в груди — цель\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 8. Руки — бицепс умеренно
                "group": "Бицепс", "exercise": "Бицепс на наклонной скамье",
                "sets": 2, "reps": "10–12", "rpe": "RIR 1", "step": "+1 кг",
                "rest": "60 сек",
                "desc": (
                    "🔥 БИЦЕПС НА НАКЛОННОЙ СКАМЬЕ\n"
                    "Категория: руки\n"
                    "Цель: умеренно — руки уже хорошо отвечают\n\n"
                    "Техника:\n"
                    "• Угол скамьи 45–60°, руки свисают вниз\n"
                    "• Полная амплитуда — растяжение в нижней точке\n"
                    "• Без читинга корпусом\n"
                    "• Медленный эксцентрик 2–3 сек\n\n"
                    "RIR 1 | Прогрессия: +1 кг"
                )
            },
            {
                # 9. Руки — трицепс поддержка
                "group": "Трицепс", "exercise": "Трицепс (канат)",
                "sets": 2, "reps": "10–12", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60 сек",
                "desc": (
                    "🔥 ТРИЦЕПС (КАНАТ)\n"
                    "Категория: руки\n"
                    "Цель: поддержка объёма рук\n\n"
                    "Техника:\n"
                    "• Локти зафиксированы у тела\n"
                    "• Полное разгибание в нижней точке\n"
                    "• Медленный возврат — не теряй напряжение\n\n"
                    "Совет: делай суперсетом с бицепсом (без отдыха между ними)\n"
                    "RIR 1 | Прогрессия: следующий стек"
                )
            },
            {
                # 10. Кор — антиротация, талия без утолщения
                "group": "Кор", "exercise": "Pallof Press",
                "sets": 3, "reps": "12 на сторону", "rpe": "RIR 1–2", "step": "+1–2 повт",
                "rest": "45–60 сек",
                "desc": (
                    "🔥 PALLOF PRESS\n"
                    "Категория: кор\n"
                    "Цель: антиротация — кор без утолщения талии\n\n"
                    "Техника:\n"
                    "• Блок сбоку на уровне груди\n"
                    "• Рукоять у груди, жми прямо вперёд и держи 1–2 сек\n"
                    "• Корпус не разворачивается — сопротивляйся вращению\n"
                    "• Ноги на ширине плеч, колени чуть согнуты\n"
                    "• 12 повторов на каждую сторону\n\n"
                    "RIR 1–2 | Прогрессия: +1–2 повтора"
                )
            },
        ],
    },

    # ─── FULL BODY C — ПАМП + АТЛЕТИК ──────────────────────────────────────────
    # День пампа и детализации верха: плечи, задняя дельта, верх спины, грудь
    # Это НЕ день ног
    "C": {
        "title": "FULL BODY C — ПАМП + АТЛЕТИК ⚡",
        "notes": [
            "Акцент: pump / athletic / верх спины и плечи",
            "Отдых: 60–90 сек",
        ],
        "review_weeks": 10,
        "exercises": [
            {
                # 1. Горизонтальный жим — памп без системной усталости
                "group": "Грудь", "exercise": "Chest Press",
                "sets": 3, "reps": "10–12", "rpe": "RIR 1–2", "step": "след. стек",
                "rest": "75–90 сек",
                "desc": (
                    "⚡ CHEST PRESS (тренажёр)\n"
                    "Категория: горизонтальный жим\n"
                    "Цель: памп груди без сильной системной усталости\n\n"
                    "Техника:\n"
                    "• Спина прижата к спинке, хват нейтральный\n"
                    "• Жим вперёд до полного разгибания\n"
                    "• Медленный возврат — не теряй напряжение\n"
                    "• Акцент на ощущение в груди, не на вес\n\n"
                    "RIR 1–2: контроль важнее веса\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 2. Вертикальная тяга — широчайшие, контроль, без перегруза
                "group": "Спина", "exercise": "Пуловер в кроссовере",
                "sets": 3, "reps": "12–15", "rpe": "RIR 1–2", "step": "след. стек",
                "rest": "75 сек",
                "desc": (
                    "⚡ ПУЛОВЕР В КРОССОВЕРЕ\n"
                    "Категория: вертикальная тяга\n"
                    "Цель: широчайшие, контроль, без перегруза\n\n"
                    "Альтернатива: тяга прямыми руками (straight-arm pulldown)\n\n"
                    "Техника:\n"
                    "• Блок высоко, трос прямой или с канатом\n"
                    "• Стоя лицом к блоку, руки прямые, хват сверху\n"
                    "• Тяни дугой вниз — к бёдрам\n"
                    "• Руки остаются прямыми всё движение\n"
                    "• Акцент на растяжение широчайших наверху\n\n"
                    "RIR 1–2: памп в широчайших — цель\n"
                    "Прогрессия: следующий стек"
                )
            },
            {
                # 3. Движение ног — фоновая нагрузка
                "group": "Ноги", "exercise": "Выпады",
                "sets": 2, "reps": "12–15", "rpe": "RIR 2", "step": "+1–2 кг",
                "rest": "75–90 сек",
                "desc": (
                    "⚡ ВЫПАДЫ\n"
                    "Категория: движение ног\n"
                    "Цель: ноги в фоне — не убиваться!\n\n"
                    "Альтернатива: лёгкий жим ногами (12–15 повторов)\n\n"
                    "Техника:\n"
                    "• Шаг вперёд, оба колена ~90°\n"
                    "• Заднее колено почти касается пола\n"
                    "• Корпус прямой, не наклоняй вперёд\n"
                    "• Чередуй ноги\n\n"
                    "RIR 2: лёгко! Это поддержка, не тренировка ног\n"
                    "Прогрессия: +1–2 кг"
                )
            },
            {
                # 4. Тазовое движение — поддержка ягодиц
                "group": "Ягодицы", "exercise": "Ягодичный мост",
                "sets": 2, "reps": "10–12", "rpe": "RIR 2", "step": "+2.5–5 кг",
                "rest": "90 сек",
                "desc": (
                    "⚡ ЯГОДИЧНЫЙ МОСТ\n"
                    "Категория: тазовое движение\n"
                    "Цель: поддержка ягодиц\n\n"
                    "Альтернатива: Hip Thrust\n\n"
                    "Техника:\n"
                    "• Лёжа на спине, ноги согнуты, стопы на полу\n"
                    "• Штанга / гиря на бёдрах\n"
                    "• Поднимай таз до прямой линии бедро-корпус\n"
                    "• Пик: сожми ягодицы, задержи 1 сек\n"
                    "• Медленно опускай\n\n"
                    "RIR 2: поддержка — без убийства\n"
                    "Прогрессия: +2.5–5 кг"
                )
            },
            {
                # 5. Горизонтальная тяга / верх спины — задняя дельта, осанка
                "group": "Задняя дельта", "exercise": "Reverse Pec Deck",
                "sets": 3, "reps": "12–15", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60–75 сек",
                "desc": (
                    "⚡ REVERSE PEC DECK\n"
                    "Категория: верх спины / горизонтальная тяга\n"
                    "Цель: задняя дельта, верх спины, осанка\n\n"
                    "Альтернатива: широкая тяга к груди (face pull широким хватом)\n\n"
                    "Техника:\n"
                    "• Сидя лицом к тренажёру, рукояти на уровне плеч\n"
                    "• Разводи руки назад — акцент на заднюю дельту\n"
                    "• Руки слегка согнуты в локтях\n"
                    "• Пик сокращения — задержка 1 сек\n"
                    "• Медленный возврат\n\n"
                    "RIR 1 | Прогрессия: следующий стек"
                )
            },
            {
                # 6. Плечи — главный памп-блок
                "group": "Плечи", "exercise": "Mechanical Drop Set (разведения)",
                "sets": 3, "reps": "12 строгих / 10 читинг / 15 частичных", "rpe": "RIR 0–1", "step": "качество",
                "rest": "90 сек",
                "desc": (
                    "⚡ MECHANICAL DROP SET — РАЗВЕДЕНИЯ\n"
                    "Категория: плечи\n"
                    "Цель: главный памп-блок дня\n\n"
                    "3 фазы без паузы внутри раунда:\n"
                    "1️⃣ 12 строгих — чистая техника, полная амплитуда\n"
                    "2️⃣ 10 читинг — небольшой импульс корпусом\n"
                    "3️⃣ 15 частичных — малая амплитуда, жжение\n\n"
                    "Всё это = 1 раунд без отдыха внутри\n"
                    "RIR 0–1: должно жечь!\n"
                    "Отдых между раундами: 90 сек\n"
                    "Прогрессия: улучшение качества"
                )
            },
            {
                # 7. Задняя дельта — добивка
                "group": "Задняя дельта", "exercise": "Face Pull",
                "sets": 3, "reps": "15–20", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60 сек",
                "desc": (
                    "⚡ FACE PULL\n"
                    "Категория: задняя дельта\n"
                    "Цель: добивка задней дельты, здоровье плеч\n\n"
                    "Техника:\n"
                    "• Блок на уровне лица или выше\n"
                    "• Тяни к лицу, локти выше плеч — в стороны\n"
                    "• Внешняя ротация в конечной точке\n"
                    "• Медленно возвращай\n\n"
                    "RIR 1 | Прогрессия: следующий стек"
                )
            },
            {
                # 8. Руки — памп бицепс
                "group": "Бицепс", "exercise": "Сгибания на бицепс",
                "sets": 2, "reps": "10–12", "rpe": "RIR 1", "step": "+1 кг",
                "rest": "60 сек",
                "desc": (
                    "⚡ СГИБАНИЯ НА БИЦЕПС\n"
                    "Категория: руки\n"
                    "Цель: памп\n\n"
                    "Варианты: гантели стоя, блок, молотки\n\n"
                    "Техника:\n"
                    "• Локти фиксированы у тела\n"
                    "• Полная амплитуда — растяжение внизу\n"
                    "• Пик сокращения наверху — задержка 1 сек\n\n"
                    "RIR 1 | Прогрессия: +1 кг"
                )
            },
            {
                # 9. Руки — памп трицепс
                "group": "Трицепс", "exercise": "Трицепс (канат)",
                "sets": 2, "reps": "10–12", "rpe": "RIR 1", "step": "след. стек",
                "rest": "60 сек",
                "desc": (
                    "⚡ ТРИЦЕПС (КАНАТ)\n"
                    "Категория: руки\n"
                    "Цель: памп\n\n"
                    "Техника:\n"
                    "• Локти зафиксированы у тела\n"
                    "• Полное разгибание в нижней точке\n"
                    "• Медленный возврат\n\n"
                    "Совет: делай суперсетом с бицепсом\n"
                    "RIR 1 | Прогрессия: следующий стек"
                )
            },
            {
                # 10. Кор — пресс
                "group": "Пресс", "exercise": "Подъём коленей в висе",
                "sets": 3, "reps": "12–15", "rpe": "RIR 1–2", "step": "+1–2 повт",
                "rest": "60 сек",
                "desc": (
                    "⚡ ПОДЪЁМ КОЛЕНЕЙ В ВИСЕ\n"
                    "Категория: кор\n"
                    "Цель: нижний пресс\n\n"
                    "Альтернатива: скручивания с весом\n\n"
                    "Техника:\n"
                    "• Вис на перекладине, тело стабильно\n"
                    "• Медленно поднимай колени к груди 2–3 сек\n"
                    "• Задержка 1 сек наверху\n"
                    "• Медленно опускай 2–3 сек\n"
                    "• Без раскачки — только пресс\n\n"
                    "RIR 1–2 | Прогрессия: +1–2 повтора"
                )
            },
        ],
    },
}

EXERCISE_CATALOG = sorted(set(
    [ex["exercise"] for d in PROGRAMS.values() for ex in d["exercises"]] +
    ["Подтягивания", "Тяга штанги в наклоне", "Face Pull", "Reverse Pec Deck",
     "Жим гантелей на наклонной скамье", "Тяга горизонтального блока", "Жим ногами",
     "Болгарские выпады", "Разведения гантелей в стороны", "Cable Lateral Raise",
     "Скручивания с весом", "Планка", "Внешняя ротация резиной", "Y-Raise",
     "Подъём ног в висе", "Cable Crunch", "Подтягивания с весом",
     "Шраги со штангой", "Икры стоя", "Икры сидя", "Cable Fly (low-to-high)",
     "Румынская тяга", "Сгибание ног лёжа", "Ягодичный мост",
     "Chest Press", "Тяга верхнего блока", "Тяга верхнего блока нейтральным хватом",
     "Mechanical Drop Set (разведения)", "Pallof Press", "Пуловер в кроссовере",
     "Бицепс на наклонной скамье", "Трицепс (канат)", "Сгибания на бицепс",
     "Подъём коленей в висе", "Выпады", "Pull-through в кроссовере",
     "Одноручная тяга в кроссовере", "Тяга с упором грудью (Chest-supported row)"]
))
PAGE_SIZE = 8

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def column_exists(cur, table_name, column_name):
    cur.execute(f"PRAGMA table_info({table_name})")
    return column_name in [row[1] for row in cur.fetchall()]

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS workouts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            workout_date TEXT NOT NULL,
            day_type TEXT NOT NULL,
            exercise TEXT NOT NULL,
            set1_reps REAL, set1_kg REAL,
            set2_reps REAL, set2_kg REAL,
            set3_reps REAL, set3_kg REAL,
            set4_reps REAL, set4_kg REAL,
            set5_reps REAL, set5_kg REAL,
            target_sets INTEGER,
            target_reps TEXT,
            target_guidance TEXT,
            step_rule TEXT,
            rpe REAL,
            notes TEXT,
            suggestion TEXT,
            coach_status TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    extra_cols = {
        "target_sets": "INTEGER",
        "target_reps": "TEXT",
        "target_guidance": "TEXT",
        "step_rule": "TEXT",
        "suggestion": "TEXT",
        "coach_status": "TEXT",
    }
    for col, typ in extra_cols.items():
        if not column_exists(cur, "workouts", col):
            cur.execute(f"ALTER TABLE workouts ADD COLUMN {col} {typ}")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS measurements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            measure_date TEXT NOT NULL,
            weight_kg REAL,
            body_fat_pct REAL,
            waist_cm REAL,
            chest_cm REAL,
            hips_cm REAL,
            arms_cm REAL,
            notes TEXT,
            photo_file_id TEXT,
            source TEXT DEFAULT 'manual',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # migration: add new columns if missing
    for col, typ in [
        ("body_fat_pct",       "REAL"),
        ("source",             "TEXT"),
        ("water_pct",          "REAL"),   # % воды в организме
        ("muscle_mass_kg",     "REAL"),   # мышечная масса (кг)
        ("fat_mass_kg",        "REAL"),   # масса жира (кг)
        ("visceral_fat_level", "INTEGER"),# уровень висцерального жира
        ("protein_pct",        "REAL"),   # % белка
        ("minerals_kg",        "REAL"),   # масса минералов (кг)
        ("biological_age",     "INTEGER"),# биологический возраст
        ("bmi",                "REAL"),   # ИМТ
        ("skeletal_muscle_kg", "REAL"),   # масса скелетных мышц
    ]:
        if not column_exists(cur, "measurements", col):
            cur.execute(f"ALTER TABLE measurements ADD COLUMN {col} {typ}")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS nutrition (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            nutrition_date TEXT NOT NULL,
            kcal INTEGER,
            protein_g REAL,
            carbs_g REAL,
            fat_g REAL,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def parse_date(s):
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s.strip(), fmt).date().isoformat()
        except ValueError:
            pass
    raise ValueError("Используй дату в формате ДД.ММ.ГГГГ")

def parse_optional_float(s):
    s = s.strip().replace(",", ".")
    if s in ("", "-", "none", "null"):
        return None
    return float(s)

def parse_sets(text):
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if not 1 <= len(parts) <= 5:
        raise ValueError("Нужно указать от 1 до 5 подходов.")
    result = []
    for part in parts:
        normalized = part.strip().replace("х", "x").replace("Х", "x").replace("×", "x").replace("*", "x")
        normalized = re.sub(r"\s+", "", normalized)
        if "x" not in normalized:
            raise ValueError("Формат: 8x60, 8х60, 8 * 60 или 8 x 60")
        reps_str, kg_str = normalized.split("x", 1)
        if not reps_str or not kg_str:
            raise ValueError("Формат: 8x60, 8х60, 8 * 60 или 8 x 60")
        result.append((parse_optional_float(reps_str), parse_optional_float(kg_str)))
    while len(result) < 5:
        result.append((None, None))
    return result

def build_repeated_sets(weight, reps, n_sets):
    result = [(reps, weight) for _ in range(max(1, min(5, n_sets)))]
    while len(result) < 5:
        result.append((None, None))
    return result

def calc_top_weight_and_reps(row):
    pairs = [(row["set1_reps"], row["set1_kg"]), (row["set2_reps"], row["set2_kg"]), (row["set3_reps"], row["set3_kg"]), (row["set4_reps"], row["set4_kg"]), (row["set5_reps"], row["set5_kg"])]
    valid = [(r, w) for r, w in pairs if r is not None and w is not None]
    if not valid:
        return None, None
    top_weight = max(w for r, w in valid)
    reps_at_top = max(r for r, w in valid if w == top_weight)
    return top_weight, reps_at_top

def calc_e1rm(top_weight, reps_at_top):
    if top_weight is None or reps_at_top is None:
        return None
    return round(float(top_weight) * (1 + float(reps_at_top) / 30), 1)

def calc_volume(row):
    total = 0.0
    for reps_col, kg_col in [("set1_reps","set1_kg"),("set2_reps","set2_kg"),("set3_reps","set3_kg"),("set4_reps","set4_kg"),("set5_reps","set5_kg")]:
        reps = row[reps_col]
        kg = row[kg_col]
        if reps is not None and kg is not None:
            total += float(reps) * float(kg)
    return round(total, 1)

def get_prev_best_e1rm(user_id, exercise, current_id=None):
    conn = get_conn()
    cur = conn.cursor()
    if current_id is None:
        cur.execute("SELECT * FROM workouts WHERE user_id=? AND exercise=? ORDER BY workout_date, id", (user_id, exercise))
    else:
        cur.execute("SELECT * FROM workouts WHERE user_id=? AND exercise=? AND id < ? ORDER BY workout_date, id", (user_id, exercise, current_id))
    rows = cur.fetchall()
    conn.close()
    vals = []
    for r in rows:
        tw, rr = calc_top_weight_and_reps(r)
        e1 = calc_e1rm(tw, rr)
        if e1 is not None:
            vals.append(e1)
    return max(vals) if vals else None

def get_last_same_exercise(user_id, exercise):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM workouts WHERE user_id=? AND exercise=? ORDER BY workout_date DESC, id DESC LIMIT 1", (user_id, exercise))
    row = cur.fetchone()
    conn.close()
    return row

def get_last_workout_date(user_id, day_type):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT workout_date FROM workouts WHERE user_id=? AND day_type=? ORDER BY workout_date DESC, id DESC LIMIT 1", (user_id, day_type))
    row = cur.fetchone()
    conn.close()
    return row["workout_date"] if row else None

def get_session_rows(user_id, day_type, workout_date):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM workouts WHERE user_id=? AND day_type=? AND workout_date=? ORDER BY id", (user_id, day_type, workout_date))
    rows = cur.fetchall()
    conn.close()
    return rows

def parse_rep_range(rep_text):
    txt = (rep_text or "-").replace(" ", "").replace("–", "-")
    if any(x in txt for x in ["сек", "sec", "/", "строгих", "читинг", "частичных"]) or txt == "-":
        return None, None
    if "-" in txt:
        a, b = txt.split("-", 1)
        try:
            return int(a), int(b)
        except ValueError:
            return None, None
    try:
        return int(txt), int(txt)
    except ValueError:
        return None, None

def find_exercise_config(day, exercise_name):
    for ex in PROGRAMS.get(day, {}).get("exercises", []):
        if ex["exercise"] == exercise_name:
            return ex
    return {"group": "Каталог", "exercise": exercise_name, "sets": None, "reps": "-", "rpe": "-", "step": "ручное решение"}

def format_last_performance(row):
    if not row:
        return "Нет прошлой записи."
    sets = []
    for i in range(1, 6):
        reps = row[f"set{i}_reps"]
        kg = row[f"set{i}_kg"]
        if reps is not None and kg is not None:
            sets.append(f"{reps:g}x{kg:g}")
    tw, rr = calc_top_weight_and_reps(row)
    e1 = calc_e1rm(tw, rr)
    return f"Прошлый раз: {', '.join(sets) if sets else '-'} | e1RM: {e1 if e1 else '-'}"

def analyze_progress(cfg, sets, rpe):
    low, high = parse_rep_range(cfg.get("reps"))
    valid = [(r, w) for r, w in sets if r is not None and w is not None]
    if not valid:
        return "Нет данных для подсказки.", "unknown"
    weights = [w for _, w in valid]
    top_weight = max(weights)
    reps_at_top = [r for r, w in valid if w == top_weight]
    min_reps_at_top = min(reps_at_top) if reps_at_top else None
    if low is None or high is None:
        if rpe is not None and rpe >= 9:
            return "Перегруз. Оставь или немного снизь вес.", "overload"
        if rpe is not None and rpe <= 7:
            return f"Легко. Можно ускорить прогрессию по правилу: {cfg['step']}.", "easy"
        return f"Свободное упражнение. Следующая цель: {cfg['step']}.", "stable"
    if min_reps_at_top is not None and min_reps_at_top >= high and (rpe is None or rpe <= 8.5):
        return f"Ты в верхней границе → пора добавлять вес. Правило: {cfg['step']}.", "progress"
    if min_reps_at_top is not None and min_reps_at_top < low:
        return "Ниже диапазона → вес пока рано увеличивать, добери повторы/технику.", "below"
    if rpe is not None and rpe >= 9:
        return "Высокий RPE → перегруз, оставь или немного снизь вес.", "overload"
    if rpe is not None and rpe <= 7:
        return "Легко → можно ускорить прогрессию.", "easy"
    return "Середина диапазона → оставь вес и добери повторы.", "stable"

def build_main_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Тренировка A 💪", callback_data="start_A")],
        [InlineKeyboardButton("Тренировка B 🔥", callback_data="start_B")],
        [InlineKeyboardButton("Тренировка C ⚡", callback_data="start_C")],
        [InlineKeyboardButton("📊 Dashboard", callback_data="dashboard_open"), InlineKeyboardButton("📥 Экспорт Excel", callback_data="export_open")],
        [InlineKeyboardButton("📏 Замеры и фото", callback_data="measure_open"), InlineKeyboardButton("🍽 Питание / ккал", callback_data="nutrition_open")],
        [InlineKeyboardButton("🤖 Спросить AI-тренера", callback_data="ai_chat")],
    ])

def build_day_menu(day, done_exercises=None, context_user_data=None):
    done = set(done_exercises or [])
    rows = []
    ex_map = {}
    for idx, ex in enumerate(PROGRAMS.get(day, {}).get("exercises", [])):
        ex_map[idx] = ex["exercise"]
        label = ("✅ " if ex["exercise"] in done else "") + ex["exercise"]
        rows.append([
            InlineKeyboardButton(label, callback_data=f"pick::{idx}"),
            InlineKeyboardButton("ℹ️", callback_data=f"exinfo::{idx}"),
        ])
    if context_user_data is not None:
        context_user_data["day_ex_map"] = ex_map
    rows.append([InlineKeyboardButton("Каталог всех упражнений", callback_data="catalog::0")])
    rows.append([InlineKeyboardButton("Добавить своё упражнение", callback_data="custom_exercise")])
    rows.append([InlineKeyboardButton("✏️ Редактировать упражнение", callback_data="edit_exercise")])
    rows.append([InlineKeyboardButton("📊 Итог тренировки / Выйти", callback_data="finish_workout")])
    rows.append([InlineKeyboardButton("В меню", callback_data="go_menu")])
    return InlineKeyboardMarkup(rows)

def get_done_exercises(user_id, day, workout_date):
    rows = get_session_rows(user_id, day, workout_date)
    return set(r["exercise"] for r in rows)

def build_catalog_menu(page, context_user_data=None):
    start = page * PAGE_SIZE
    items = EXERCISE_CATALOG[start:start + PAGE_SIZE]
    rows = []
    cat_map = {}
    for local_idx, name in enumerate(items):
        global_idx = start + local_idx
        cat_map[global_idx] = name
        rows.append([InlineKeyboardButton(name, callback_data=f"cat_pick::{global_idx}")])
    if context_user_data is not None:
        existing = context_user_data.get("cat_ex_map", {})
        existing.update(cat_map)
        context_user_data["cat_ex_map"] = existing
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ Назад", callback_data=f"catalog::{page-1}"))
    if start + PAGE_SIZE < len(EXERCISE_CATALOG):
        nav.append(InlineKeyboardButton("Вперед ▶️", callback_data=f"catalog::{page+1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton("К упражнениям дня", callback_data="back_to_day_menu")])
    rows.append([InlineKeyboardButton("Добавить своё упражнение", callback_data="custom_exercise")])
    rows.append([InlineKeyboardButton("📊 Итог тренировки / Выйти", callback_data="finish_workout")])
    return InlineKeyboardMarkup(rows)

def build_input_mode_menu(has_last):
    rows = []
    if has_last:
        rows.append([InlineKeyboardButton("Повторить прошлый вес", callback_data="use_last")])
    rows.append([InlineKeyboardButton("Быстрый ввод", callback_data="quick_input")])
    rows.append([InlineKeyboardButton("Ввести вручную", callback_data="manual_input")])
    rows.append([InlineKeyboardButton("Назад к упражнениям", callback_data="back_to_day_menu")])
    return InlineKeyboardMarkup(rows)

def build_edit_select_menu(session_rows):
    """Inline keyboard: pick which exercise from current session to edit."""
    rows = []
    for r in session_rows:
        sets_str = ", ".join(
            f"{r[f'set{i}_reps']:g}x{r[f'set{i}_kg']:g}"
            for i in range(1, 6)
            if r[f"set{i}_reps"] is not None and r[f"set{i}_kg"] is not None
        )
        label = f"{r['exercise']} [{sets_str}]"
        rows.append([InlineKeyboardButton(label, callback_data=f"edit_pick::{r['id']}")])
    rows.append([InlineKeyboardButton("Назад к упражнениям", callback_data="back_to_day_menu")])
    return InlineKeyboardMarkup(rows)

def build_edit_field_menu(row_id):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("Изменить подходы (повторения×кг)", callback_data=f"edit_sets::{row_id}")],
        [InlineKeyboardButton("Изменить заметку", callback_data=f"edit_notes::{row_id}")],
        [InlineKeyboardButton("Назад к списку", callback_data="edit_exercise")],
    ])


def build_workout_summary(user_id, day, workout_date):
    rows = get_session_rows(user_id, day, workout_date)
    if not rows:
        return f"🏁 Тренировка {day} | {workout_date}\nНет записей."

    total_volume = round(sum(calc_volume(r) for r in rows), 1)
    progress_list = []
    overload_list = []
    scored = []

    for r in rows:
        tw, rr = calc_top_weight_and_reps(r)
        e1 = calc_e1rm(tw, rr)
        scored.append((r["exercise"], e1 if e1 is not None else 0, calc_volume(r)))
        status = r["coach_status"] or ""
        if status == "progress":
            progress_list.append(r["exercise"])
        elif status == "overload":
            overload_list.append(r["exercise"])

    best_by_volume = sorted(scored, key=lambda x: x[2], reverse=True)[:3]

    lines = [
        f"🏁 Тренировка {day} | {workout_date} завершена!",
        f"Упражнений: {len(rows)} | Тоннаж: {total_volume} кг",
        "",
        "Топ по тоннажу:",
    ]
    for ex, _, vol in best_by_volume:
        lines.append(f"• {ex}: {vol:g} кг")

    if progress_list:
        lines.append("")
        lines.append("Прогресс:")
        for ex in progress_list[:5]:
            lines.append(f"• {ex}")

    if overload_list:
        lines.append("")
        lines.append("Перегруз:")
        for ex in overload_list[:5]:
            lines.append(f"• {ex}")

    return "\n".join(lines)


async def start(update, context):
    text = "Привет. Это v6.1.\n\nТеперь после завершения тренировки бот выводит итог: тоннаж, лучшие упражнения, где был прогресс и где был перегруз."
    target = update.message if update.message else update.callback_query.message
    await target.reply_text(text, reply_markup=build_main_menu())

async def help_cmd(update, context):
    await update.message.reply_text(
        "Сценарий:\n1) Нажимаешь Тренировка A/B/C\n2) Вводишь дату\n3) Выбираешь упражнение\n4) Можно повторить прошлый вес, сделать быстрый ввод или ввести вручную\n5) После упражнения бот даёт авто-подсказку\n6) После завершения тренировки бот выдаёт итог\n7) /coach — недельный анализ"
    )

async def cancel(update, context):
    context.user_data.clear()
    target = update.message if update.message else update.callback_query.message
    if update.callback_query:
        await update.callback_query.answer()
    await target.reply_text("Ок, отменил.", reply_markup=ReplyKeyboardRemove())
    await target.reply_text("Меню:", reply_markup=build_main_menu())
    return ConversationHandler.END

async def menu_callback(update, context):
    q = update.callback_query
    await q.answer()
    data = q.data
    if data.startswith("start_"):
        day = data.split("_", 1)[1]
        context.user_data.clear()
        context.user_data["selected_day"] = day
        last_dt = get_last_workout_date(update.effective_user.id, day)
        review_msg = "📌 Контроль пересмотра тренировок: через 10 недель"
        if last_dt:
            try:
                last_date = datetime.fromisoformat(last_dt).date()
                review_msg = f"📌 Контроль пересмотра тренировок: {(last_date + timedelta(weeks=10)).isoformat()} (через 10 недель от последней {day})"
            except Exception:
                pass
        notes = "\n".join(f"• {x}" for x in PROGRAMS[day]["notes"])
        await q.message.reply_text(
            f"{PROGRAMS[day]['title']}\n\nРекомендации:\n{notes}\n\n{review_msg}\n\nДата тренировки? Формат ДД.ММ.ГГГГ",
            reply_markup=ReplyKeyboardRemove(),
        )
        return SESSION_DATE
    if data == "dashboard_open":
        await dashboard_text(q.message, update.effective_user.id)
        await q.message.reply_text("Меню:", reply_markup=build_main_menu())
        return ConversationHandler.END
    if data == "export_open":
        await export_file(q.message, update.effective_user.id)
        await q.message.reply_text("Меню:", reply_markup=build_main_menu())
        return ConversationHandler.END
    if data == "finish_workout":
        day = context.user_data.get("selected_day", "-")
        workout_date = context.user_data.get("workout_date", "-")
        summary = build_workout_summary(update.effective_user.id, day, workout_date)
        user_id = update.effective_user.id
        context.user_data.clear()
        await q.message.reply_text(summary, reply_markup=build_main_menu())
        # AI-анализ (отдельным сообщением после итога)
        await q.message.chat.send_action("typing")
        ai_text = await ai_analyze_workout(user_id, day, workout_date, q.message.bot)
        if ai_text:
            await q.message.reply_text(ai_text)
        # Напоминание о замерах раз в месяц
        if check_measurement_reminder(user_id):
            await q.message.reply_text(
                "📏 Напоминание: прошёл месяц с последних замеров!\n"
                "Не забудь внести замеры и фото — это важно для отслеживания прогресса.",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Внести замеры сейчас", callback_data="measure_open")]
                ])
            )
        return ConversationHandler.END
    if data == "go_menu":
        context.user_data.clear()
        await q.message.reply_text("Меню:", reply_markup=build_main_menu())
        return ConversationHandler.END
    if data == "back_to_day_menu":
        day = context.user_data.get("selected_day")
        if not day:
            await q.message.reply_text("Сначала начни тренировку.", reply_markup=build_main_menu())
            return ConversationHandler.END
        workout_date = context.user_data.get("workout_date", "")
        done = get_done_exercises(update.effective_user.id, day, workout_date)
        await q.message.reply_text(f"Упражнения дня {day}:", reply_markup=build_day_menu(day, done, context.user_data))
        return SESSION_MANUAL_INPUT
    if data.startswith("catalog::"):
        page = int(data.split("::", 1)[1])
        await q.message.reply_text("Каталог упражнений:", reply_markup=build_catalog_menu(page, context.user_data))
        return SESSION_MANUAL_INPUT
    if data == "custom_exercise":
        await q.message.reply_text("Напиши название своего упражнения:")
        return CUSTOM_EXERCISE_NAME
    if data.startswith("pick::"):
        idx = int(data.split("::", 1)[1])
        ex_map = context.user_data.get("day_ex_map", {})
        ex_name = ex_map.get(idx)
        if not ex_name:
            # fallback: rebuild map from PROGRAMS
            day = context.user_data.get("selected_day", "")
            exercises = PROGRAMS.get(day, {}).get("exercises", [])
            ex_name = exercises[idx]["exercise"] if idx < len(exercises) else None
        if not ex_name:
            await q.message.reply_text("Ошибка: упражнение не найдено.")
            return SESSION_MANUAL_INPUT
        context.user_data["current_exercise"] = ex_name
        day = context.user_data.get("selected_day", "")
        cfg = find_exercise_config(day, ex_name)
        prev_row = get_last_same_exercise(update.effective_user.id, ex_name)
        await q.message.reply_text(
            f"Выбрано: {ex_name}\nГруппа: {cfg['group']}\nЦель: {cfg['sets'] if cfg['sets'] else '-'} подходов × {cfg['reps']}\nОриентир: {cfg['rpe']}\nШаг прогрессии: {cfg['step']}\n{format_last_performance(prev_row)}\n\nКак хочешь занести данные?",
            reply_markup=build_input_mode_menu(prev_row is not None),
        )
        return SESSION_MANUAL_INPUT
    if data.startswith("cat_pick::"):
        idx = int(data.split("::", 1)[1])
        cat_map = context.user_data.get("cat_ex_map", {})
        ex_name = cat_map.get(idx)
        if not ex_name:
            ex_name = EXERCISE_CATALOG[idx] if idx < len(EXERCISE_CATALOG) else None
        if not ex_name:
            await q.message.reply_text("Ошибка: упражнение не найдено.")
            return SESSION_MANUAL_INPUT
        context.user_data["current_exercise"] = ex_name
        day = context.user_data.get("selected_day", "")
        cfg = find_exercise_config(day, ex_name)
        prev_row = get_last_same_exercise(update.effective_user.id, ex_name)
        await q.message.reply_text(
            f"Выбрано: {ex_name}\nГруппа: {cfg['group']}\nЦель: {cfg['sets'] if cfg['sets'] else '-'} подходов × {cfg['reps']}\nОриентир: {cfg['rpe']}\nШаг прогрессии: {cfg['step']}\n{format_last_performance(prev_row)}\n\nКак хочешь занести данные?",
            reply_markup=build_input_mode_menu(prev_row is not None),
        )
        return SESSION_MANUAL_INPUT
    if data == "use_last":
        ex_name = context.user_data.get("current_exercise")
        if not ex_name:
            await q.message.reply_text("Сначала выбери упражнение.")
            return ConversationHandler.END
        prev_row = get_last_same_exercise(update.effective_user.id, ex_name)
        if not prev_row:
            await q.message.reply_text("Нет прошлой записи. Используй быстрый ввод или ручной.")
            return SESSION_MANUAL_INPUT
        sets = []
        for i in range(1, 6):
            reps = prev_row[f"set{i}_reps"]
            kg = prev_row[f"set{i}_kg"]
            if reps is not None and kg is not None:
                sets.append((reps, kg))
        while len(sets) < 5:
            sets.append((None, None))
        context.user_data["current_sets"] = sets[:5]
        await q.message.reply_text(f"Подставил прошлый вариант: {format_last_performance(prev_row)}\nТеперь введи RPE или '-'")
        return SESSION_RPE
    if data == "quick_input":
        ex_name = context.user_data.get("current_exercise")
        if not ex_name:
            await q.message.reply_text("Сначала выбери упражнение.")
            return ConversationHandler.END
        day = context.user_data.get("selected_day", "")
        cfg = find_exercise_config(day, ex_name)
        default_sets = cfg["sets"] or 4
        await q.message.reply_text(
            f"Быстрый ввод в формате:\nвес повторения количество_подходов\nНапример: 60 8 {default_sets}\n\nЕсли количество подходов не укажешь, бот возьмёт целевое число подходов."
        )
        return SESSION_QUICK_INPUT
    if data == "manual_input":
        await q.message.reply_text("Введи подходы вручную. Пример:\n8x60, 8x60, 7x62.5, 6x62.5")
        return SESSION_MANUAL_INPUT
    if data == "edit_exercise":
        day = context.user_data.get("selected_day")
        workout_date = context.user_data.get("workout_date", "")
        if not day or not workout_date:
            await q.message.reply_text("Нет активной тренировки.")
            return ConversationHandler.END
        session_rows = get_session_rows(update.effective_user.id, day, workout_date)
        if not session_rows:
            await q.message.reply_text("В этой тренировке ещё нет записей для редактирования.")
            return SESSION_MANUAL_INPUT
        await q.message.reply_text("Выбери упражнение для редактирования:", reply_markup=build_edit_select_menu(session_rows))
        return SESSION_MANUAL_INPUT
    if data.startswith("edit_pick::"):
        row_id = int(data.split("::", 1)[1])
        context.user_data["edit_row_id"] = row_id
        await q.message.reply_text("Что хочешь изменить?", reply_markup=build_edit_field_menu(row_id))
        return SESSION_MANUAL_INPUT
    if data.startswith("edit_sets::"):
        row_id = int(data.split("::", 1)[1])
        context.user_data["edit_row_id"] = row_id
        context.user_data["edit_field"] = "sets"
        await q.message.reply_text(
            "Введи новые подходы в формате:\n8x60, 8x60, 7x62.5\n\nМожно указать от 1 до 5 подходов."
        )
        return SESSION_EDIT
    if data.startswith("edit_notes::"):
        row_id = int(data.split("::", 1)[1])
        context.user_data["edit_row_id"] = row_id
        context.user_data["edit_field"] = "notes"
        await q.message.reply_text("Введи новую заметку (или '-' чтобы очистить):")
        return SESSION_EDIT
    return ConversationHandler.END

async def session_date(update, context):
    try:
        context.user_data["workout_date"] = parse_date(update.message.text)
    except Exception as e:
        await update.message.reply_text(str(e))
        return SESSION_DATE
    day = context.user_data["selected_day"]
    done = get_done_exercises(update.effective_user.id, day, context.user_data["workout_date"])
    await update.message.reply_text(
        f"Готово. День {day} | дата: {context.user_data['workout_date']}\nТеперь выбирай упражнения в любом порядке:",
        reply_markup=build_day_menu(day, done, context.user_data),
    )
    return SESSION_MANUAL_INPUT

async def custom_exercise_name(update, context):
    ex_name = update.message.text.strip()
    if not ex_name:
        await update.message.reply_text("Название не может быть пустым.")
        return CUSTOM_EXERCISE_NAME
    context.user_data["current_exercise"] = ex_name
    prev_row = get_last_same_exercise(update.effective_user.id, ex_name)
    await update.message.reply_text(
        f"Добавлено своё упражнение: {ex_name}\n{format_last_performance(prev_row)}\n\nКак хочешь занести данные?",
        reply_markup=build_input_mode_menu(prev_row is not None),
    )
    return SESSION_MANUAL_INPUT

async def session_quick_input(update, context):
    ex_name = context.user_data.get("current_exercise")
    if not ex_name:
        await update.message.reply_text("Сначала выбери упражнение.")
        return ConversationHandler.END
    parts = update.message.text.strip().replace(",", ".").split()
    if len(parts) not in {2, 3}:
        await update.message.reply_text("Формат быстрого ввода: вес повторения количество_подходов\nПример: 60 8 4")
        return SESSION_QUICK_INPUT
    try:
        weight = float(parts[0])
        reps = float(parts[1])
        n_sets = int(float(parts[2])) if len(parts) == 3 else (find_exercise_config(context.user_data.get("selected_day", ""), ex_name)["sets"] or 4)
    except Exception:
        await update.message.reply_text("Не понял числа. Пример: 60 8 4")
        return SESSION_QUICK_INPUT
    context.user_data["current_sets"] = build_repeated_sets(weight, reps, n_sets)
    await update.message.reply_text(f"Ок, собрал {n_sets} подход(а/ов) по шаблону: {reps:g}x{weight:g}\nТеперь введи RPE или '-'")
    return SESSION_RPE

async def session_manual_input(update, context):
    if "current_exercise" not in context.user_data:
        day = context.user_data.get("selected_day")
        if day:
            workout_date = context.user_data.get("workout_date", "")
            done = get_done_exercises(update.effective_user.id, day, workout_date)
            await update.message.reply_text("Сначала выбери упражнение кнопкой ниже:", reply_markup=build_day_menu(day, done, context.user_data))
            return SESSION_MANUAL_INPUT
        await update.message.reply_text("Сначала начни тренировку.", reply_markup=build_main_menu())
        return ConversationHandler.END
    try:
        context.user_data["current_sets"] = parse_sets(update.message.text)
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}\n\nПримеры: 15x40, 8х67.5, 6 * 67.5, 5 x 67.5")
        return SESSION_MANUAL_INPUT
    await update.message.reply_text("RPE? Если не используешь — '-'")
    return SESSION_RPE

async def session_rpe(update, context):
    try:
        context.user_data["current_rpe"] = parse_optional_float(update.message.text)
    except Exception:
        await update.message.reply_text("RPE должен быть числом или '-'.")
        return SESSION_RPE
    await update.message.reply_text("Заметка по упражнению? Если не нужна — '-'")
    return SESSION_NOTES

async def session_notes(update, context):
    user_id = update.effective_user.id
    ex_name = context.user_data["current_exercise"]
    day = context.user_data.get("selected_day", "")
    workout_date = context.user_data.get("workout_date", "")
    cfg = find_exercise_config(day, ex_name)
    notes = update.message.text.strip()
    notes = None if notes == "-" else notes
    sets = context.user_data["current_sets"]
    rpe = context.user_data["current_rpe"]
    suggestion, coach_status = analyze_progress(cfg, sets, rpe)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO workouts (user_id, workout_date, day_type, exercise, set1_reps, set1_kg, set2_reps, set2_kg, set3_reps, set3_kg, set4_reps, set4_kg, set5_reps, set5_kg, target_sets, target_reps, target_guidance, step_rule, rpe, notes, suggestion, coach_status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, workout_date, day, ex_name, sets[0][0], sets[0][1], sets[1][0], sets[1][1], sets[2][0], sets[2][1], sets[3][0], sets[3][1], sets[4][0], sets[4][1], cfg["sets"], cfg["reps"], cfg["rpe"], cfg["step"], rpe, notes, suggestion, coach_status)
    )
    row_id = cur.lastrowid
    conn.commit()
    cur.execute("SELECT * FROM workouts WHERE id=?", (row_id,))
    row = cur.fetchone()
    conn.close()

    tw, rr = calc_top_weight_and_reps(row)
    e1rm = calc_e1rm(tw, rr)
    prev_best = get_prev_best_e1rm(user_id, ex_name, row_id)
    pr = "да" if (e1rm is not None and (prev_best is None or e1rm > prev_best)) else "нет"
    delta = None if (e1rm is None or prev_best is None) else round(e1rm - prev_best, 1)

    for key in ["current_exercise", "current_sets", "current_rpe"]:
        context.user_data.pop(key, None)

    done = get_done_exercises(user_id, day, workout_date)
    await update.message.reply_text(
        f"Сохранил: {ex_name} ✅\ne1RM: {e1rm if e1rm is not None else '-'}\nPR: {pr}\nΔ к прошлому лучшему: {delta if delta is not None else '-'}\n\nПодсказка тренера:\n{suggestion}\n\nВыбирай следующее упражнение:",
        reply_markup=build_day_menu(day, done, context.user_data),
    )
    return SESSION_MANUAL_INPUT

async def session_edit(update, context):
    """Handle text input when editing an existing exercise record."""
    user_id = update.effective_user.id
    row_id = context.user_data.get("edit_row_id")
    field = context.user_data.get("edit_field")
    day = context.user_data.get("selected_day", "")
    workout_date = context.user_data.get("workout_date", "")

    if not row_id or not field:
        await update.message.reply_text("Ошибка: нет данных для редактирования.")
        return ConversationHandler.END

    conn = get_conn()
    cur = conn.cursor()

    if field == "sets":
        try:
            sets = parse_sets(update.message.text)
        except Exception as e:
            await update.message.reply_text(f"Ошибка: {e}\n\nПримеры: 15x40, 8х67.5, 6 * 67.5, 5 x 67.5")
            return SESSION_EDIT
        cur.execute(
            """UPDATE workouts SET
               set1_reps=?, set1_kg=?,
               set2_reps=?, set2_kg=?,
               set3_reps=?, set3_kg=?,
               set4_reps=?, set4_kg=?,
               set5_reps=?, set5_kg=?
               WHERE id=? AND user_id=?""",
            (sets[0][0], sets[0][1], sets[1][0], sets[1][1],
             sets[2][0], sets[2][1], sets[3][0], sets[3][1],
             sets[4][0], sets[4][1], row_id, user_id)
        )
        conn.commit()
        # Recalculate suggestion/coach_status with updated sets
        cur.execute("SELECT * FROM workouts WHERE id=?", (row_id,))
        row = cur.fetchone()
        if row:
            cfg = find_exercise_config(day, row["exercise"])
            suggestion, coach_status = analyze_progress(cfg, sets, row["rpe"])
            cur.execute(
                "UPDATE workouts SET suggestion=?, coach_status=? WHERE id=?",
                (suggestion, coach_status, row_id)
            )
            conn.commit()
        conn.close()
        context.user_data.pop("edit_row_id", None)
        context.user_data.pop("edit_field", None)
        done = get_done_exercises(user_id, day, workout_date)
        await update.message.reply_text(
            "✅ Подходы обновлены.\n\nВыбирай следующее упражнение:",
            reply_markup=build_day_menu(day, done, context.user_data)
        )
        return SESSION_MANUAL_INPUT

    elif field == "notes":
        text = update.message.text.strip()
        new_notes = None if text == "-" else text
        cur.execute("UPDATE workouts SET notes=? WHERE id=? AND user_id=?", (new_notes, row_id, user_id))
        conn.commit()
        conn.close()
        context.user_data.pop("edit_row_id", None)
        context.user_data.pop("edit_field", None)
        done = get_done_exercises(user_id, day, workout_date)
        await update.message.reply_text(
            "✅ Заметка обновлена.\n\nВыбирай следующее упражнение:",
            reply_markup=build_day_menu(day, done, context.user_data)
        )
        return SESSION_MANUAL_INPUT

    conn.close()
    return SESSION_MANUAL_INPUT

async def dashboard_text(message, user_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM workouts WHERE user_id=?", (user_id,))
    rows = cur.fetchall()
    conn.close()

    workouts_cnt = len(set((r["workout_date"], r["day_type"]) for r in rows))
    total_volume = round(sum(calc_volume(r) for r in rows), 1)
    best_by_ex = {}
    for r in rows:
        tw, rr = calc_top_weight_and_reps(r)
        e1 = calc_e1rm(tw, rr)
        if e1 is not None:
            best_by_ex[r["exercise"]] = max(best_by_ex.get(r["exercise"], 0), e1)
    lines = ["📊 Dashboard", f"Тренировок: {workouts_cnt}", f"Тоннаж: {total_volume} кг", f"Упражнений с прогрессом: {len(best_by_ex)}"]
    if best_by_ex:
        lines += ["", "Лучшие e1RM:"]
        for ex, e1 in sorted(best_by_ex.items(), key=lambda x: x[1], reverse=True)[:8]:
            lines.append(f"• {ex}: {e1}")
    await message.reply_text("\n".join(lines))

async def dashboard(update, context):
    await dashboard_text(update.message, update.effective_user.id)

async def coach(update, context):
    user_id = update.effective_user.id
    cutoff = (datetime.utcnow().date() - timedelta(days=7)).isoformat()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM workouts WHERE user_id=? AND workout_date>=? ORDER BY workout_date, id", (user_id, cutoff))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await update.message.reply_text("За последние 7 дней записей пока нет.")
        return

    workouts_cnt = len(set((r["workout_date"], r["day_type"]) for r in rows))
    total_volume = round(sum(calc_volume(r) for r in rows), 1)
    status_counts = {"progress": 0, "stable": 0, "below": 0, "overload": 0, "easy": 0, "unknown": 0}
    for r in rows:
        status = r["coach_status"] or "unknown"
        if status not in status_counts:
            status = "unknown"
        status_counts[status] += 1

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM workouts WHERE user_id=? ORDER BY workout_date, id", (user_id,))
    all_rows = cur.fetchall()
    conn.close()

    by_ex = {}
    for r in all_rows:
        tw, rr = calc_top_weight_and_reps(r)
        e1 = calc_e1rm(tw, rr)
        if e1 is None:
            continue
        by_ex.setdefault(r["exercise"], []).append((r["workout_date"], e1))

    progress_lines = []
    regress = 0
    stagnation = 0
    improvements = 0
    for ex, vals in by_ex.items():
        if len(vals) >= 2:
            prev = vals[-2][1]
            last = vals[-1][1]
            diff = round(last - prev, 1)
            if diff > 0:
                improvements += 1
                progress_lines.append(f"+ {ex} → +{diff} кг e1RM")
            elif diff < 0:
                regress += 1
                progress_lines.append(f"- {ex} → {diff} кг e1RM")
            else:
                stagnation += 1
                progress_lines.append(f"= {ex} → без изменений")

    total_tracked = improvements + regress + stagnation
    balance_line = "Недостаточно данных."
    if total_tracked > 0:
        p = round(improvements / total_tracked * 100)
        s = round(stagnation / total_tracked * 100)
        r = round(regress / total_tracked * 100)
        balance_line = f"Баланс: {p}% прогресс / {s}% плато / {r}% регресс"

    recommendation = "Продолжаем."
    if status_counts["overload"] >= max(2, workouts_cnt):
        recommendation = "⚠️ Есть признаки перегруза: высокий RPE и/или просадка повторений. Оставь веса на 1–2 тренировки или чуть снизь объём."
    elif status_counts["easy"] >= max(2, workouts_cnt):
        recommendation = "⚠️ Есть признаки недогруза: упражнения ощущаются слишком легко. Можно ускорить прогрессию."
    elif regress > improvements:
        recommendation = "⚠️ Регресса больше, чем роста. Проверь восстановление, сон и объём."
    elif improvements >= regress and improvements > 0:
        recommendation = "✅ Вектор хороший. Большинство движений идут в нужную сторону."

    msg = [
        "🧠 Coach — анализ за последние 7 дней",
        f"Тренировок: {workouts_cnt}",
        f"Общий тоннаж: {total_volume} кг",
        "",
        "Статусы:",
        f"• прогресс: {status_counts['progress']}",
        f"• стабильно: {status_counts['stable']}",
        f"• ниже диапазона: {status_counts['below']}",
        f"• перегруз: {status_counts['overload']}",
        f"• легко: {status_counts['easy']}",
        "",
        balance_line,
        "",
        "Тренд по упражнениям:",
    ]
    msg.extend(progress_lines[:12] if progress_lines else ["Пока мало данных по повторным записям."])
    msg.extend(["", recommendation])
    await update.message.reply_text("\n".join(msg))

def create_excel_export(user_id, output_path):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM workouts WHERE user_id=? ORDER BY workout_date, id", (user_id,))
    workouts = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    headers = ["Date", "Week", "Day", "Exercise", "Set1 Reps", "Set1 Kg", "Set2 Reps", "Set2 Kg", "Set3 Reps", "Set3 Kg", "Set4 Reps", "Set4 Kg", "Set5 Reps", "Set5 Kg", "Target Sets", "Target Reps", "Target Guidance", "Step Rule", "RPE", "Volume", "Top Weight", "Top Reps@Top", "e1RM", "PR", "Suggestion", "Coach Status", "Notes"]
    ws.append(headers)

    navy = PatternFill("solid", fgColor="1F4E78")
    a_fill = PatternFill("solid", fgColor="DCEEFF")
    b_fill = PatternFill("solid", fgColor="E6F4DE")
    c_fill = PatternFill("solid", fgColor="FFF2D8")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    for c in ws[1]:
        c.fill = navy
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    first_date = None
    best_so_far = {}
    for r in workouts:
        if first_date is None:
            first_date = datetime.fromisoformat(r["workout_date"]).date()
        week = ((datetime.fromisoformat(r["workout_date"]).date() - first_date).days // 7) + 1
        tw, rr = calc_top_weight_and_reps(r)
        e1 = calc_e1rm(tw, rr)
        prev = best_so_far.get(r["exercise"])
        is_pr = e1 is not None and (prev is None or e1 > prev)
        if is_pr and e1 is not None:
            best_so_far[r["exercise"]] = e1
        ws.append([r["workout_date"], week, r["day_type"], r["exercise"], r["set1_reps"], r["set1_kg"], r["set2_reps"], r["set2_kg"], r["set3_reps"], r["set3_kg"], r["set4_reps"], r["set4_kg"], r["set5_reps"], r["set5_kg"], r["target_sets"], r["target_reps"], r["target_guidance"], r["step_rule"], r["rpe"], calc_volume(r), tw, rr, e1, "PR" if is_pr else "", r["suggestion"], r["coach_status"], r["notes"]])

    for row in range(2, ws.max_row + 1):
        day = ws.cell(row=row, column=3).value
        fill = a_fill if day == "A" else b_fill if day == "B" else c_fill if day == "C" else None
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
        if ws.cell(row=row, column=24).value == "PR":
            ws.cell(row=row, column=24).fill = yellow

    widths = [12,8,6,30,10,9,10,9,10,9,10,9,10,9,10,16,16,14,8,12,11,12,10,7,28,14,24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i) if i <= 26 else "AA"].width = w
    ws.freeze_panes = "A2"

    prog = wb.create_sheet("Progress")
    prog.append(["Date", "Жим штанги лёжа", "Присед со штангой", "Подтягивания с весом", "Жим гантелей на наклонной скамье"])
    exercise_map = {"Жим штанги лёжа": 2, "Присед со штангой": 3, "Подтягивания с весом": 4, "Жим гантелей на наклонной скамье": 5}
    by_date = {}
    for r in workouts:
        tw, rr = calc_top_weight_and_reps(r)
        e1 = calc_e1rm(tw, rr)
        if e1 is None:
            continue
        dt = r["workout_date"]
        by_date.setdefault(dt, {})
        col = exercise_map.get(r["exercise"])
        if col:
            prev = by_date[dt].get(col)
            by_date[dt][col] = max(prev, e1) if prev is not None else e1
    for dt in sorted(by_date):
        row = [dt, None, None, None, None]
        for col, val in by_date[dt].items():
            row[col - 1] = val
        prog.append(row)
    for c in prog[1]:
        c.fill = navy
        c.font = Font(color="FFFFFF", bold=True)
    chart = LineChart()
    chart.title = "Strength Progress (e1RM)"
    chart.y_axis.title = "kg"
    chart.x_axis.title = "Date"
    data = Reference(prog, min_col=2, max_col=5, min_row=1, max_row=prog.max_row)
    cats = Reference(prog, min_col=1, min_row=2, max_row=prog.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    prog.add_chart(chart, "G2")
    wb.save(output_path)
    return output_path

async def export_file(message, user_id):
    path = f"training_export_v61_{user_id}.xlsx"
    create_excel_export(user_id, path)
    with open(path, "rb") as f:
        await message.reply_document(document=f, filename=path, caption="Вот твоя выгрузка Excel 📊")

async def export_cmd(update, context):
    await export_file(update.message, update.effective_user.id)

def build_ai_context(user_id: int) -> str:
    """Собирает последние тренировки из БД в текст для системного промпта."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM workouts WHERE user_id=? ORDER BY workout_date DESC, id DESC LIMIT 60",
        (user_id,)
    )
    rows = cur.fetchall()
    conn.close()
    if not rows:
        return "История тренировок пуста."

    # группируем по дате+дню
    sessions = {}
    for r in rows:
        key = (r["workout_date"], r["day_type"])
        sessions.setdefault(key, []).append(r)

    lines = []
    for (date, day), exercises in sorted(sessions.items(), reverse=True):
        lines.append(f"\nТренировка {day} | {date}")
        for r in exercises:
            sets = ", ".join(
                f"{r[f'set{i}_reps']:g}x{r[f'set{i}_kg']:g}"
                for i in range(1, 6)
                if r[f"set{i}_reps"] is not None and r[f"set{i}_kg"] is not None
            )
            rpe_str = f" | RPE {r['rpe']:g}" if r["rpe"] is not None else ""
            note_str = f" | Заметка: {r['notes']}" if r["notes"] else ""
            status_str = f" | Статус: {r['coach_status']}" if r["coach_status"] else ""
            lines.append(f"  • {r['exercise']}: {sets}{rpe_str}{status_str}{note_str}")
    return "\n".join(lines)


async def ai_chat_start(update, context):
    """AI-чат временно отключён."""
    # CLAUDE_API_DISABLED — раскомментировать весь блок когда будет готово
    q = update.callback_query
    await q.answer()
    await q.message.reply_text(
        "🤖 AI-тренер временно отключён. Скоро будет доступен.",
        reply_markup=build_main_menu()
    )
    return ConversationHandler.END


async def ai_chat_message(update, context):
    # CLAUDE_API_DISABLED
    return ConversationHandler.END


async def ai_chat_stop(update, context):
    """Выход из режима AI-чата."""
    context.user_data.pop("ai_history", None)
    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_text("Вышел из чата с AI.", reply_markup=build_main_menu())
    else:
        await update.message.reply_text("Вышел из чата с AI.", reply_markup=build_main_menu())
    return ConversationHandler.END


# ─── ℹ️ ОПИСАНИЕ УПРАЖНЕНИЯ ────────────────────────────────────────────────

async def exercise_info_callback(update, context):
    q = update.callback_query
    await q.answer()
    idx = int(q.data.split("::", 1)[1])
    day = context.user_data.get("selected_day", "")
    exercises = PROGRAMS.get(day, {}).get("exercises", [])
    if idx >= len(exercises):
        await q.message.reply_text("Описание не найдено.")
        return SESSION_MANUAL_INPUT
    ex = exercises[idx]
    desc = ex.get("desc") or (
        f"{ex['exercise']}\n"
        f"Группа: {ex['group']}\n"
        f"Подходы×повторы: {ex.get('sets','?')}×{ex['reps']}\n"
        f"RPE: {ex['rpe']}\n"
        f"Отдых: {ex.get('rest','—')}\n"
        f"Прогрессия: {ex['step']}"
    )
    await q.message.reply_text(desc)
    return SESSION_MANUAL_INPUT


# ─── AI-АНАЛИЗ ТРЕНИРОВКИ (ВРЕМЕННО ОТКЛЮЧЁН) ─────────────────────────────

async def ai_analyze_workout(user_id: int, day: str, workout_date: str, bot) -> str:
    # CLAUDE_API_DISABLED — раскомментировать когда API будет подключён
    return ""


# ─── ЕЖЕДНЕВНЫЕ НАПОМИНАНИЯ И НЕДЕЛЬНАЯ СТАТИСТИКА ─────────────────────────

ADMIN_USER_ID = int(os.getenv("ADMIN_USER_ID", "0"))


async def send_daily_weight_reminder(context):
    """Ежедневное напоминание взвеситься в 07:00 (Буэнос-Айрес)."""
    if not ADMIN_USER_ID:
        return
    await context.bot.send_message(
        chat_id=ADMIN_USER_ID,
        text=(
            "⚖️ <b>Доброе утро! Время взвеситься.</b>\n\n"
            "Введи данные весов командой /measure\n"
            "или нажми кнопку ниже 👇\n\n"
            "Формат быстрого ввода:\n"
            "<code>вес жир% вода% мышцы_кг</code>\n"
            "Пример: <code>64.8 17.0 59.0 51.0</code>"
        ),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⚖️ Ввести замеры", callback_data="measure_open")]
        ])
    )


async def send_weekly_weight_report(context):
    """Еженедельный отчёт по весу — каждый понедельник в 08:00."""
    if not ADMIN_USER_ID:
        return
    conn = get_conn()
    cur = conn.cursor()
    today = datetime.now().date()
    week_start = (today - timedelta(days=7)).isoformat()
    prev_week_start = (today - timedelta(days=14)).isoformat()

    cur.execute("""
        SELECT AVG(weight_kg) as avg_w, AVG(body_fat_pct) as avg_fat,
               MIN(weight_kg) as min_w, MAX(weight_kg) as max_w, COUNT(*) as n
        FROM measurements
        WHERE user_id=? AND measure_date >= ? AND weight_kg IS NOT NULL
    """, (ADMIN_USER_ID, week_start))
    this_week = cur.fetchone()

    cur.execute("""
        SELECT AVG(weight_kg) as avg_w, AVG(body_fat_pct) as avg_fat
        FROM measurements
        WHERE user_id=? AND measure_date >= ? AND measure_date < ? AND weight_kg IS NOT NULL
    """, (ADMIN_USER_ID, prev_week_start, week_start))
    last_week = cur.fetchone()
    conn.close()

    if not this_week or not this_week["n"]:
        await context.bot.send_message(
            chat_id=ADMIN_USER_ID,
            text="📊 <b>Недельный отчёт</b>\n\nНет данных за прошедшую неделю. Взвешивайся каждое утро!",
            parse_mode="HTML"
        )
        return

    avg_w = round(this_week["avg_w"], 2)
    avg_fat = round(this_week["avg_fat"], 1) if this_week["avg_fat"] else "—"

    trend_w = ""
    if last_week and last_week["avg_w"]:
        diff = round(avg_w - last_week["avg_w"], 2)
        arrow = "📉" if diff < 0 else ("📈" if diff > 0 else "➡️")
        trend_w = f"{arrow} {'+' if diff > 0 else ''}{diff} кг vs прошлой недели"

    trend_fat = ""
    if last_week and last_week["avg_fat"] and this_week["avg_fat"]:
        diff_fat = round(this_week["avg_fat"] - last_week["avg_fat"], 1)
        trend_fat = f"\n  Динамика жира: {'+' if diff_fat > 0 else ''}{diff_fat}%"

    msg = (
        f"📊 <b>Недельный отчёт — вес</b>\n\n"
        f"Средний вес: <b>{avg_w} кг</b>\n"
        f"  Разброс: {this_week['min_w']}–{this_week['max_w']} кг\n"
        f"  Средний жир: <b>{avg_fat}%</b>\n"
        f"  Взвешиваний: {this_week['n']}/7\n\n"
        f"{trend_w}{trend_fat}\n\n"
        f"{'✅ Взвешивался каждый день!' if this_week['n'] == 7 else f'⚠️ Пропущено {7 - this_week[chr(110)]} дней — взвешивайся ежедневно для точного тренда.'}"
    )
    await context.bot.send_message(chat_id=ADMIN_USER_ID, text=msg, parse_mode="HTML")


# ─── ЗАМЕРЫ И ФОТО ─────────────────────────────────────────────────────────

def get_last_measurement(user_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM measurements WHERE user_id=? ORDER BY measure_date DESC LIMIT 1",
        (user_id,)
    )
    row = cur.fetchone()
    conn.close()
    return row

def check_measurement_reminder(user_id):
    """Возвращает True если последний замер был больше 28 дней назад или его нет."""
    last = get_last_measurement(user_id)
    if not last:
        return True
    last_date = datetime.fromisoformat(last["measure_date"]).date()
    return (datetime.now().date() - last_date).days >= 28

async def measure_open_callback(update, context):
    """Главное подменю замеров."""
    q = update.callback_query
    await q.answer()
    last = get_last_measurement(update.effective_user.id)
    if last:
        fat_str = f" | Жир: {last['body_fat_pct']}%" if last.get('body_fat_pct') else ""
        waist_str = f"\n  Талия: {last['waist_cm']} см" if last.get('waist_cm') else ""
        chest_str = f" | Грудь: {last['chest_cm']} см" if last.get('chest_cm') else ""
        hips_str  = f" | Бёдра: {last['hips_cm']} см" if last.get('hips_cm') else ""
        arms_str  = f" | Руки: {last['arms_cm']} см" if last.get('arms_cm') else ""
        src = last.get('source') or 'manual'
        prev = (
            f"Последний замер: {last['measure_date']} ({src})\n"
            f"  Вес: {last['weight_kg']} кг{fat_str}"
            f"{waist_str}{chest_str}{hips_str}{arms_str}\n"
        )
    else:
        prev = "Замеров пока нет.\n"

    await q.message.reply_text(
        f"📏 ЗАМЕРЫ И ФОТО\n\n{prev}",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✍️ Ввести замеры вручную", callback_data="measure_manual")],
            [InlineKeyboardButton("⚖️ Импорт CSV умных весов", callback_data="measure_scale_csv")],
            [InlineKeyboardButton("🍎 Импорт Apple Health XML", callback_data="measure_health_xml")],
            [InlineKeyboardButton("📊 История замеров", callback_data="measure_history")],
            [InlineKeyboardButton("← В меню", callback_data="go_menu")],
        ])
    )
    return MEASURE_INPUT


async def measure_manual_callback(update, context):
    q = update.callback_query
    await q.answer()
    await q.message.reply_text(
        "⚖️ <b>Введи данные с умных весов</b>\n\n"
        "Формат (через пробел):\n"
        "<code>вес жир% вода% мышцы_кг</code>\n\n"
        "Примеры:\n"
        "<code>64.8 17.0 59.0 51.0</code> — основные параметры\n"
        "<code>64.8 17.0</code> — только вес и жир\n"
        "<code>64.8</code> — только вес\n\n"
        "Расширенный формат (все параметры весов):\n"
        "<code>вес жир% вода% мышцы_кг висцер_жир белок% минералы_кг биол_возраст</code>\n"
        "Пример: <code>64.8 17.0 59.0 51.0 5 18.8 2.8 29</code>\n\n"
        "Прочерк <code>-</code> пропускает параметр",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Отмена", callback_data="measure_open")]
        ])
    )
    context.user_data["measure_mode"] = "manual"
    return MEASURE_INPUT


async def measure_scale_csv_callback(update, context):
    """Инструкция по импорту CSV с умных весов."""
    q = update.callback_query
    await q.answer()
    await q.message.reply_text(
        "⚖️ <b>ИМПОРТ CSV — Xiaomi / Huawei / Renpho</b>\n\n"
        "<b>Как экспортировать из Renpho:</b>\n"
        "Настройки → Данные → Экспорт данных → выбери период → CSV\n\n"
        "<b>Как экспортировать из Xiaomi Mi Fit / Zepp:</b>\n"
        "Профиль → Конфиденциальность → Экспорт данных\n\n"
        "Ожидаемые колонки (хотя бы одна из):\n"
        "<code>Date, Weight, Body Fat, BMI</code>\n\n"
        "Отправь CSV-файл 👇",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Отмена", callback_data="measure_open")]
        ])
    )
    context.user_data["measure_mode"] = "scale_csv"
    return IMPORT_SCALE


async def measure_health_xml_callback(update, context):
    """Инструкция по импорту Apple Health XML."""
    q = update.callback_query
    await q.answer()
    await q.message.reply_text(
        "🍎 <b>ИМПОРТ APPLE HEALTH XML</b>\n\n"
        "<b>Как экспортировать:</b>\n"
        "1. Приложение «Здоровье» → твой профиль (аватар)\n"
        "2. «Экспортировать данные о здоровье» → ZIP\n"
        "3. Распакуй ZIP → найди <code>export.xml</code>\n"
        "4. Отправь этот XML-файл сюда\n\n"
        "Бот импортирует: вес тела, % жира, ккал (активные)\n\n"
        "Отправь файл <code>export.xml</code> 👇",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Отмена", callback_data="measure_open")]
        ])
    )
    context.user_data["measure_mode"] = "health_xml"
    return IMPORT_HEALTH


async def measure_history_callback(update, context):
    """Показывает последние 10 замеров."""
    q = update.callback_query
    await q.answer()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM measurements WHERE user_id=? ORDER BY measure_date DESC LIMIT 10",
        (update.effective_user.id,)
    )
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await q.message.reply_text("Замеров пока нет.", reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("← Назад", callback_data="measure_open")]
        ]))
        return MEASURE_INPUT

    lines = ["📊 <b>История замеров:</b>\n"]
    for r in rows:
        fat = f" | {r['body_fat_pct']}%" if r.get('body_fat_pct') else ""
        w = f"{r['weight_kg']} кг" if r.get('weight_kg') else "—"
        lines.append(f"<b>{r['measure_date']}</b>: {w}{fat}")
        parts = []
        if r.get('waist_cm'): parts.append(f"талия {r['waist_cm']}")
        if r.get('chest_cm'): parts.append(f"грудь {r['chest_cm']}")
        if r.get('hips_cm'):  parts.append(f"бёдра {r['hips_cm']}")
        if r.get('arms_cm'):  parts.append(f"руки {r['arms_cm']}")
        if parts:
            lines.append("  " + " | ".join(parts) + " см")
    await q.message.reply_text(
        "\n".join(lines), parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("← Назад", callback_data="measure_open")]
        ])
    )
    return MEASURE_INPUT


async def measure_input(update, context):
    """Ручной ввод замеров текстом — расширенный формат для умных весов."""
    user_id = update.effective_user.id
    text = update.message.text.strip()
    parts = text.split()

    def pf(v):
        v = v.strip().replace(",", ".")
        return None if v in ("-", "") else float(v)
    def pi(v):
        v = v.strip().replace(",", ".")
        return None if v in ("-", "") else int(float(v))

    try:
        weight          = pf(parts[0]) if len(parts) > 0 else None
        fat_pct         = pf(parts[1]) if len(parts) > 1 else None
        water_pct       = pf(parts[2]) if len(parts) > 2 else None
        muscle_mass_kg  = pf(parts[3]) if len(parts) > 3 else None
        visceral_fat    = pi(parts[4]) if len(parts) > 4 else None
        protein_pct     = pf(parts[5]) if len(parts) > 5 else None
        minerals_kg     = pf(parts[6]) if len(parts) > 6 else None
        biological_age  = pi(parts[7]) if len(parts) > 7 else None
        # Авторасчёт жировой массы
        fat_mass_kg = round(weight * fat_pct / 100, 2) if weight and fat_pct else None
        # ИМТ не вводим вручную — берётся с весов отдельно
    except (ValueError, IndexError) as e:
        await update.message.reply_text(f"Ошибка: {e}\nПример: 64.8 17.0 59.0 51.0")
        return MEASURE_INPUT

    context.user_data["pending_measure"] = {
        "weight_kg": weight, "body_fat_pct": fat_pct,
        "water_pct": water_pct, "muscle_mass_kg": muscle_mass_kg,
        "fat_mass_kg": fat_mass_kg, "visceral_fat_level": visceral_fat,
        "protein_pct": protein_pct, "minerals_kg": minerals_kg,
        "biological_age": biological_age,
        "measure_date": datetime.now().date().isoformat(),
        "source": "manual",
    }

    # Быстрое подтверждение
    lines = [f"✅ <b>Данные за {context.user_data['pending_measure']['measure_date']}:</b>"]
    if weight:         lines.append(f"  Вес: {weight} кг")
    if fat_pct:        lines.append(f"  Жир: {fat_pct}% ({fat_mass_kg} кг)")
    if water_pct:      lines.append(f"  Вода: {water_pct}%")
    if muscle_mass_kg: lines.append(f"  Мышцы: {muscle_mass_kg} кг")
    if visceral_fat:   lines.append(f"  Висцер. жир: {visceral_fat}")
    if protein_pct:    lines.append(f"  Белок: {protein_pct}%")
    if minerals_kg:    lines.append(f"  Минералы: {minerals_kg} кг")
    if biological_age: lines.append(f"  Биол. возраст: {biological_age} лет")

    await update.message.reply_text(
        "\n".join(lines) + "\n\nОтправь фото 📸 или нажми «Сохранить».",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("💾 Сохранить", callback_data="measure_save_nophoto")]
        ])
    )
    return MEASURE_INPUT


async def import_scale_csv(update, context):
    """Импорт CSV файла с умных весов (Renpho / Xiaomi / Huawei)."""
    user_id = update.effective_user.id
    doc = update.message.document
    if not doc or not doc.file_name.lower().endswith(".csv"):
        await update.message.reply_text("Пожалуйста, отправь файл в формате .csv")
        return IMPORT_SCALE

    file = await doc.get_file()
    raw = await file.download_as_bytearray()
    text = raw.decode("utf-8", errors="replace")

    # Пробуем несколько разделителей
    dialect = "excel"
    for sep in [",", ";", "\t"]:
        if text.count(sep) > text.count(",") if sep != "," else True:
            dialect = None
            break

    reader = csv.DictReader(io.StringIO(text))
    rows_saved = 0
    skipped = 0

    # Нормализация имён колонок
    COL_MAP = {
        "date": ["date", "дата", "time", "measurement date", "weigh-in date"],
        "weight": ["weight", "вес", "weight(kg)", "weight (kg)", "body weight"],
        "fat": ["body fat", "body fat%", "fat%", "% body fat", "body fat percentage", "fat ratio"],
    }

    def find_col(headers, variants):
        h_lower = {h.lower().strip(): h for h in headers}
        for v in variants:
            if v in h_lower:
                return h_lower[v]
        return None

    conn = get_conn()
    cur = conn.cursor()

    try:
        rows = list(reader)
        if not rows:
            raise ValueError("Файл пустой")
        headers = rows[0].keys()
        date_col   = find_col(headers, COL_MAP["date"])
        weight_col = find_col(headers, COL_MAP["weight"])
        fat_col    = find_col(headers, COL_MAP["fat"])

        if not date_col:
            raise ValueError(f"Не найдена колонка даты. Колонки в файле: {', '.join(headers)}")
        if not weight_col and not fat_col:
            raise ValueError(f"Не найдены колонки веса/жира. Колонки: {', '.join(headers)}")

        for row in rows:
            try:
                raw_date = row[date_col].strip()
                # пробуем несколько форматов даты
                d = None
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y", "%Y/%m/%d", "%d/%m/%Y"):
                    try:
                        d = datetime.strptime(raw_date.split()[0], fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
                if not d:
                    skipped += 1
                    continue

                def safe_float(v):
                    if v is None: return None
                    v = str(v).strip().replace(",", ".").replace("%", "")
                    return float(v) if v else None

                weight = safe_float(row.get(weight_col)) if weight_col else None
                fat    = safe_float(row.get(fat_col)) if fat_col else None

                # Пропускаем дубликаты (та же дата + тот же источник)
                cur.execute(
                    "SELECT id FROM measurements WHERE user_id=? AND measure_date=? AND source=?",
                    (user_id, d, "scale_csv")
                )
                if cur.fetchone():
                    skipped += 1
                    continue

                cur.execute("""
                    INSERT INTO measurements (user_id, measure_date, weight_kg, body_fat_pct, source)
                    VALUES (?,?,?,?,?)
                """, (user_id, d, weight, fat, "scale_csv"))
                rows_saved += 1
            except Exception:
                skipped += 1
                continue

        conn.commit()
    except Exception as e:
        conn.close()
        await update.message.reply_text(f"❌ Ошибка при разборе файла: {e}")
        return IMPORT_SCALE
    conn.close()

    await update.message.reply_text(
        f"✅ Импорт завершён!\n"
        f"  Добавлено: {rows_saved} записей\n"
        f"  Пропущено: {skipped}\n\n"
        f"{'Колонки найдены: ' + (weight_col or '') + (' | ' + fat_col if fat_col else '')}",
        reply_markup=build_main_menu()
    )
    return ConversationHandler.END


async def import_health_xml(update, context):
    """Импорт export.xml из Apple Health."""
    user_id = update.effective_user.id
    doc = update.message.document
    if not doc or not doc.file_name.lower().endswith(".xml"):
        await update.message.reply_text("Пожалуйста, отправь файл export.xml из Apple Health.")
        return IMPORT_HEALTH

    await update.message.chat.send_action("typing")
    file = await doc.get_file()
    raw = await file.download_as_bytearray()

    weight_data = {}   # date -> kg
    fat_data = {}      # date -> %
    kcal_data = {}     # date -> kcal

    WEIGHT_ID  = "HKQuantityTypeIdentifierBodyMass"
    FAT_ID     = "HKQuantityTypeIdentifierBodyFatPercentage"
    KCAL_ID    = "HKQuantityTypeIdentifierActiveEnergyBurned"

    try:
        root = ET.fromstring(raw)
        for rec in root.iter("Record"):
            rtype = rec.get("type", "")
            raw_date = rec.get("startDate", "")[:10]
            val_str  = rec.get("value", "")
            unit     = rec.get("unit", "")
            try:
                val = float(val_str)
            except ValueError:
                continue

            if rtype == WEIGHT_ID:
                # конвертация из lbs если нужно
                if "lb" in unit.lower():
                    val = val * 0.453592
                weight_data[raw_date] = round(val, 2)
            elif rtype == FAT_ID:
                fat_data[raw_date] = round(val * 100, 1)
            elif rtype == KCAL_ID:
                kcal_data[raw_date] = kcal_data.get(raw_date, 0) + val
    except ET.ParseError as e:
        await update.message.reply_text(f"❌ Не удалось разобрать XML: {e}")
        return IMPORT_HEALTH

    conn = get_conn()
    cur = conn.cursor()
    w_saved = fat_saved = kcal_saved = skipped = 0

    all_dates = set(weight_data) | set(fat_data)
    for d in sorted(all_dates):
        w   = weight_data.get(d)
        fat = fat_data.get(d)
        cur.execute(
            "SELECT id FROM measurements WHERE user_id=? AND measure_date=? AND source=?",
            (user_id, d, "apple_health")
        )
        if cur.fetchone():
            skipped += 1
            continue
        cur.execute("""
            INSERT INTO measurements (user_id, measure_date, weight_kg, body_fat_pct, source)
            VALUES (?,?,?,?,?)
        """, (user_id, d, w, fat, "apple_health"))
        if w:   w_saved += 1
        if fat: fat_saved += 1

    # Сохраняем ккал в nutrition
    for d, total_kcal in kcal_data.items():
        cur.execute(
            "SELECT id FROM nutrition WHERE user_id=? AND nutrition_date=? AND notes=?",
            (user_id, d, "apple_health")
        )
        if cur.fetchone():
            continue
        cur.execute("""
            INSERT INTO nutrition (user_id, nutrition_date, kcal, notes)
            VALUES (?,?,?,?)
        """, (user_id, d, int(total_kcal), "apple_health"))
        kcal_saved += 1

    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"✅ Apple Health импортирован!\n\n"
        f"  Вес: {w_saved} дней\n"
        f"  % жира: {fat_saved} дней\n"
        f"  Активные ккал: {kcal_saved} дней\n"
        f"  Пропущено (уже есть): {skipped}",
        reply_markup=build_main_menu()
    )
    return ConversationHandler.END


async def measure_photo(update, context):
    user_id = update.effective_user.id
    photo = update.message.photo[-1]
    pending = context.user_data.get("pending_measure")
    if not pending:
        await update.message.reply_text("Сначала введи замеры цифрами.")
        return MEASURE_INPUT
    pending["photo_file_id"] = photo.file_id
    _save_measurement(user_id, pending)
    context.user_data.pop("pending_measure", None)
    await update.message.reply_text("✅ Замеры и фото сохранены!", reply_markup=build_main_menu())
    return ConversationHandler.END


async def measure_save_nophoto_callback(update, context):
    q = update.callback_query
    await q.answer()
    user_id = update.effective_user.id
    pending = context.user_data.get("pending_measure")
    if not pending:
        await q.message.reply_text("Нет данных для сохранения.")
        return ConversationHandler.END
    _save_measurement(user_id, pending)
    context.user_data.pop("pending_measure", None)
    await q.message.reply_text("✅ Замеры сохранены!", reply_markup=build_main_menu())
    return ConversationHandler.END


def _save_measurement(user_id, data):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO measurements
            (user_id, measure_date, weight_kg, body_fat_pct,
             waist_cm, chest_cm, hips_cm, arms_cm, photo_file_id, source,
             water_pct, muscle_mass_kg, fat_mass_kg, visceral_fat_level,
             protein_pct, minerals_kg, biological_age, bmi, skeletal_muscle_kg)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        user_id, data["measure_date"],
        data.get("weight_kg"),   data.get("body_fat_pct"),
        data.get("waist_cm"),    data.get("chest_cm"),
        data.get("hips_cm"),     data.get("arms_cm"),
        data.get("photo_file_id"), data.get("source", "manual"),
        data.get("water_pct"),   data.get("muscle_mass_kg"),
        data.get("fat_mass_kg"), data.get("visceral_fat_level"),
        data.get("protein_pct"), data.get("minerals_kg"),
        data.get("biological_age"), data.get("bmi"),
        data.get("skeletal_muscle_kg"),
    ))
    conn.commit()
    conn.close()


# ─── ПИТАНИЕ / ККАЛ ────────────────────────────────────────────────────────

def get_last_nutrition(user_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM nutrition WHERE user_id=? AND (notes IS NULL OR notes != 'apple_health') ORDER BY nutrition_date DESC LIMIT 5",
        (user_id,)
    )
    rows = cur.fetchall()
    conn.close()
    return rows


async def nutrition_open_callback(update, context):
    q = update.callback_query
    await q.answer()
    rows = get_last_nutrition(update.effective_user.id)
    if rows:
        hist = "\n".join(
            f"  {r['nutrition_date']}: {r['kcal']} ккал"
            + (f" | Б:{r['protein_g']}г" if r['protein_g'] else "")
            + (f" У:{r['carbs_g']}г" if r['carbs_g'] else "")
            + (f" Ж:{r['fat_g']}г" if r['fat_g'] else "")
            for r in rows
        )
        hist_text = f"Последние записи:\n{hist}\n\n"
    else:
        hist_text = "Записей пока нет.\n\n"

    await q.message.reply_text(
        f"🍽 ПИТАНИЕ\n\n{hist_text}",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✍️ Ввести вручную", callback_data="nutrition_manual")],
            [InlineKeyboardButton("🍎 Импорт Apple Health XML", callback_data="measure_health_xml")],
            [InlineKeyboardButton("← В меню", callback_data="go_menu")],
        ])
    )
    return NUTRITION_INPUT


async def nutrition_manual_callback(update, context):
    q = update.callback_query
    await q.answer()
    await q.message.reply_text(
        "Введи данные в формате:\n"
        "<b>ккал белки углеводы жиры</b>\n\n"
        "Примеры:\n"
        "<code>2400</code> — только калории\n"
        "<code>2400 180 220 65</code> — ккал + БУЖ\n",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Отмена", callback_data="nutrition_open")]
        ])
    )
    return NUTRITION_INPUT


async def nutrition_input(update, context):
    user_id = update.effective_user.id
    text = update.message.text.strip()
    parts = text.split()

    def pf(v):
        v = v.replace(",", ".")
        return None if v == "-" else float(v)

    try:
        kcal    = int(float(parts[0])) if len(parts) > 0 else None
        protein = pf(parts[1]) if len(parts) > 1 else None
        carbs   = pf(parts[2]) if len(parts) > 2 else None
        fat     = pf(parts[3]) if len(parts) > 3 else None
    except (ValueError, IndexError) as e:
        await update.message.reply_text(f"Ошибка: {e}\nПример: 2400 180 220 65")
        return NUTRITION_INPUT

    conn = get_conn()
    cur = conn.cursor()
    today = datetime.now().date().isoformat()
    cur.execute(
        "INSERT INTO nutrition (user_id, nutrition_date, kcal, protein_g, carbs_g, fat_g) VALUES (?,?,?,?,?,?)",
        (user_id, today, kcal, protein, carbs, fat)
    )
    conn.commit()
    conn.close()

    msg = f"✅ Сохранено: {today}\n  Ккал: {kcal}"
    if protein: msg += f" | Б: {protein}г"
    if carbs:   msg += f" | У: {carbs}г"
    if fat:     msg += f" | Ж: {fat}г"

    await update.message.reply_text(msg, reply_markup=build_main_menu())
    return ConversationHandler.END


def build_application():
    if not BOT_TOKEN:
        raise RuntimeError("Нужно задать TELEGRAM_BOT_TOKEN")
    init_db()
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # ── Ежедневное напоминание взвеситься — 07:00 Буэнос-Айрес (10:00 UTC) ──
    if ADMIN_USER_ID and app.job_queue:
        from datetime import time as dtime, timezone
        app.job_queue.run_daily(
            send_daily_weight_reminder,
            time=dtime(10, 0, tzinfo=timezone.utc),   # 07:00 Buenos Aires (UTC-3)
            name="daily_weight_reminder",
        )
        # Еженедельный отчёт — каждый понедельник 08:00 Buenos Aires (11:00 UTC)
        app.job_queue.run_daily(
            send_weekly_weight_report,
            time=dtime(11, 0, tzinfo=timezone.utc),
            days=(0,),  # 0 = понедельник
            name="weekly_weight_report",
        )
        logger.info(f"Jobs scheduled: daily reminder + weekly report for user {ADMIN_USER_ID}")

    conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(menu_callback, pattern=r"^(start_[ABC]|dashboard_open|export_open|finish_workout|go_menu|back_to_day_menu|catalog::\d+|pick::\d+|cat_pick::\d+|custom_exercise|use_last|quick_input|manual_input|edit_exercise|edit_pick::\d+|edit_sets::\d+|edit_notes::\d+)$"),
            CallbackQueryHandler(ai_chat_start, pattern=r"^ai_chat$"),
            CallbackQueryHandler(measure_open_callback, pattern=r"^measure_open$"),
            CallbackQueryHandler(nutrition_open_callback, pattern=r"^nutrition_open$"),
            CommandHandler("trainer", start),
        ],
        states={
            SESSION_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, session_date)],
            CUSTOM_EXERCISE_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, custom_exercise_name)],
            SESSION_QUICK_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, session_quick_input)],
            SESSION_MANUAL_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, session_manual_input),
                CallbackQueryHandler(menu_callback, pattern=r"^(finish_workout|go_menu|back_to_day_menu|catalog::\d+|pick::\d+|cat_pick::\d+|custom_exercise|use_last|quick_input|manual_input|edit_exercise|edit_pick::\d+|edit_sets::\d+|edit_notes::\d+)$"),
                CallbackQueryHandler(exercise_info_callback, pattern=r"^exinfo::\d+$"),
            ],
            SESSION_RPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, session_rpe)],
            SESSION_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, session_notes)],
            SESSION_EDIT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, session_edit),
                CallbackQueryHandler(menu_callback, pattern=r"^(edit_exercise|edit_pick::\d+|edit_sets::\d+|edit_notes::\d+|back_to_day_menu)$"),
            ],
            AI_CHAT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, ai_chat_message),
                CallbackQueryHandler(ai_chat_stop, pattern=r"^ai_stop$"),
                CommandHandler("stop", ai_chat_stop),
            ],
            MEASURE_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, measure_input),
                MessageHandler(filters.PHOTO, measure_photo),
                CallbackQueryHandler(measure_open_callback,        pattern=r"^measure_open$"),
                CallbackQueryHandler(measure_manual_callback,      pattern=r"^measure_manual$"),
                CallbackQueryHandler(measure_scale_csv_callback,   pattern=r"^measure_scale_csv$"),
                CallbackQueryHandler(measure_health_xml_callback,  pattern=r"^measure_health_xml$"),
                CallbackQueryHandler(measure_history_callback,     pattern=r"^measure_history$"),
                CallbackQueryHandler(measure_save_nophoto_callback,pattern=r"^measure_save_nophoto$"),
            ],
            IMPORT_SCALE: [
                MessageHandler(filters.Document.ALL, import_scale_csv),
                CallbackQueryHandler(measure_open_callback, pattern=r"^measure_open$"),
            ],
            IMPORT_HEALTH: [
                MessageHandler(filters.Document.ALL, import_health_xml),
                CallbackQueryHandler(measure_open_callback,       pattern=r"^measure_open$"),
                CallbackQueryHandler(nutrition_open_callback,     pattern=r"^nutrition_open$"),
            ],
            NUTRITION_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, nutrition_input),
                CallbackQueryHandler(nutrition_open_callback,   pattern=r"^nutrition_open$"),
                CallbackQueryHandler(nutrition_manual_callback, pattern=r"^nutrition_manual$"),
                CallbackQueryHandler(measure_health_xml_callback, pattern=r"^measure_health_xml$"),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("stop", ai_chat_stop),
            CallbackQueryHandler(measure_open_callback, pattern=r"^measure_open$"),
        ],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("dashboard", dashboard))
    app.add_handler(CommandHandler("coach", coach))
    app.add_handler(CommandHandler("export", export_cmd))
    app.add_handler(CommandHandler("stop", ai_chat_stop))
    app.add_handler(conv)
    app.add_handler(CommandHandler("cancel", cancel))
    return app

HEALTH_WEBHOOK_TOKEN = os.getenv("HEALTH_WEBHOOK_TOKEN", "")
HEALTH_WEBHOOK_PORT  = int(os.getenv("HEALTH_WEBHOOK_PORT", "8080"))


# ─── AIOHTTP WEBHOOK — приём данных от Apple Shortcuts ─────────────────────

async def handle_health_data(request: web.Request) -> web.Response:
    """
    POST /health
    Headers: X-Token: <HEALTH_WEBHOOK_TOKEN>
    JSON body:
    {
        "user_id": 87400066,
        "date": "2026-05-04",          // необязательно, по умолчанию сегодня
        "weight_kg": 82.5,             // необязательно
        "body_fat_pct": 18.3,          // необязательно
        "kcal": 2400,                  // необязательно
        "waist_cm": 78.0,              // необязательно
        "chest_cm": 98.0,              // необязательно
        "hips_cm": 96.0,               // необязательно
        "arms_cm": 36.0                // необязательно
    }
    """
    # Проверка токена
    token = request.headers.get("X-Token", "")
    if not HEALTH_WEBHOOK_TOKEN or token != HEALTH_WEBHOOK_TOKEN:
        return web.Response(status=401, text="Unauthorized")

    try:
        data = await request.json()
    except Exception:
        return web.Response(status=400, text="Invalid JSON")

    user_id = data.get("user_id")
    if not user_id:
        return web.Response(status=400, text="Missing user_id")

    measure_date = data.get("date", datetime.now().date().isoformat())
    weight   = data.get("weight_kg")
    fat      = data.get("body_fat_pct")
    kcal     = data.get("kcal")
    waist    = data.get("waist_cm")
    chest    = data.get("chest_cm")
    hips     = data.get("hips_cm")
    arms     = data.get("arms_cm")

    saved_measure = False
    saved_kcal    = False

    conn = get_conn()
    cur  = conn.cursor()

    # Сохраняем замеры если есть вес или % жира
    if weight is not None or fat is not None:
        cur.execute(
            "SELECT id FROM measurements WHERE user_id=? AND measure_date=? AND source=?",
            (user_id, measure_date, "shortcut")
        )
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO measurements
                    (user_id, measure_date, weight_kg, body_fat_pct, waist_cm, chest_cm, hips_cm, arms_cm, source)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (user_id, measure_date, weight, fat, waist, chest, hips, arms, "shortcut"))
            saved_measure = True

    # Сохраняем ккал если есть
    if kcal is not None:
        cur.execute(
            "SELECT id FROM nutrition WHERE user_id=? AND nutrition_date=? AND notes=?",
            (user_id, measure_date, "shortcut")
        )
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO nutrition (user_id, nutrition_date, kcal, notes)
                VALUES (?,?,?,?)
            """, (user_id, measure_date, int(kcal), "shortcut"))
            saved_kcal = True

    conn.commit()
    conn.close()

    # Уведомляем пользователя в Telegram
    app = request.app.get("tg_app")
    if app and (saved_measure or saved_kcal):
        parts = []
        if weight:   parts.append(f"⚖️ Вес: {weight} кг")
        if fat:      parts.append(f"💧 Жир: {fat}%")
        if kcal:     parts.append(f"🍽 Ккал: {kcal}")
        if waist:    parts.append(f"📏 Талия: {waist} см")
        msg = f"✅ Apple Health — данные за {measure_date}:\n" + "\n".join(parts)
        try:
            await app.bot.send_message(chat_id=user_id, text=msg)
        except Exception as e:
            logger.warning(f"Не удалось отправить уведомление: {e}")

    return web.json_response({
        "ok": True,
        "saved_measure": saved_measure,
        "saved_kcal": saved_kcal,
        "date": measure_date,
    })


async def handle_ping(request: web.Request) -> web.Response:
    return web.Response(text="ok")


def build_aiohttp_app(tg_app) -> web.Application:
    aio = web.Application()
    aio["tg_app"] = tg_app
    aio.router.add_get("/ping", handle_ping)
    aio.router.add_post("/health", handle_health_data)
    return aio


def main():
    tg_app = build_application()

    async def run_all():
        aio_app = build_aiohttp_app(tg_app)
        runner  = web.AppRunner(aio_app)
        await runner.setup()
        site = web.TCPSite(runner, "0.0.0.0", HEALTH_WEBHOOK_PORT)
        await site.start()
        logger.info(f"Health webhook listening on port {HEALTH_WEBHOOK_PORT}")

        async with tg_app:
            await tg_app.initialize()
            await tg_app.start()
            await tg_app.updater.start_polling()
            logger.info("Telegram bot started")
            # держим процесс живым
            stop_event = asyncio.Event()
            try:
                await stop_event.wait()
            except (KeyboardInterrupt, asyncio.CancelledError):
                pass
            finally:
                await tg_app.updater.stop()
                await tg_app.stop()
                await runner.cleanup()

    asyncio.run(run_all())


if __name__ == "__main__":
    main()
